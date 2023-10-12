import pandas as pd
from openpyxl import load_workbook, Workbook

import xlsxwriter as xls
from ..connection import connect_db

def get_value(df, strategy, value_to_fetch):
        val = df.loc[df.MARKET_LUCRATIVENESS == strategy][value_to_fetch]
        if not val.empty:
            return val.iloc[0] 
        else:
            return 0


class DistrictClassificationHelper:
    # def get_value(df, strategy, value_to_fetch):
    #     val = df.loc[df.MARKET_LUCRATIVENESS == strategy][value_to_fetch]
    #     if not val.empty:
    #         return val.iloc[0] 
    #     else:
    #         return 0
                
    def get_file(df,month):
        # df= pd.read_csv('District_monthly_classification.csv')

        data = df.groupby(['ZONE', 'STATE', 'MARKET_LUCRATIVENESS'], as_index= False).agg({'SALES':'sum', 'MARKET_POTENTIAL':'sum'})

        summed_data= data.groupby(['ZONE'], as_index= False)['SALES'].sum()
        summed_data.rename(columns={'SALES': 'TOTAL_SALES'}, inplace= True)

        final_df = data.merge(summed_data, on='ZONE')
        final_df['MARKET SHARE'] = final_df['SALES'] / final_df['MARKET_POTENTIAL']
        final_df['SALES_PROPORTION'] = final_df['SALES'] / final_df['TOTAL_SALES']
        final_df.drop(columns=['TOTAL_SALES'], inplace= True)

        grand_total = final_df.groupby(['ZONE', 'STATE'], as_index= False).sum(numeric_only= True)
        grand_total['MARKET_LUCRATIVENESS'] = 'Grand Total'
        final_df = final_df.append(grand_total, ignore_index= True)
        # -----------------------------------------------------------------------------------------
        # zone level sum
        all_zones = df.groupby(['ZONE', 'MARKET_LUCRATIVENESS'], as_index= False).agg({'SALES':'sum', 'MARKET_POTENTIAL':'sum'})
        all_zones['MARKET SHARE'] = all_zones['SALES'] / all_zones['MARKET_POTENTIAL']
        all_zones['TOTAL_SALES']= all_zones['SALES'].sum()
        all_zones['SALES_PROPORTION'] = all_zones['SALES'] / all_zones['TOTAL_SALES']
        all_zones.drop(columns=['TOTAL_SALES'], inplace= True)
        all_zones = all_zones.append(all_zones.groupby(['ZONE'], as_index= False).sum(numeric_only=True), ignore_index=True)
        all_zones.fillna('Grand Total', inplace= True)

        # -----------------------------------------------------------------------------------------
            
            
        all_india = df.groupby(['MARKET_LUCRATIVENESS'], as_index= False).agg({'SALES':'sum', 'MARKET_POTENTIAL':'sum'})
        all_india['MARKET SHARE'] = all_india['SALES'] / all_india['MARKET_POTENTIAL']
        all_india['TOTAL_SALES']= all_india['SALES'].sum()
        all_india['SALES_PROPORTION'] = all_india['SALES'] / all_india['TOTAL_SALES']
        all_india.drop(columns=['TOTAL_SALES'], inplace= True)
        all_india = all_india.append(all_india.sum(numeric_only=True), ignore_index=True)
        all_india.iloc[4, 0]= 'Grand Total'
        # -----------------------------------------------------------------------------------------

        file_path = 'analytical_data/district_cls_data/district_cls_out_65.xlsx'

        wb = Workbook()
        # ws = wb.get_active_sheet()
        # ws = wb.create_sheet()

        # wb = load_workbook(file_path)
        # wb = xls.Workbook(file_path)
        # ws = wb.get_worksheet_by_name(name='Sheet2')
        ws = wb['Sheet']


        ws.cell(1, 1).value = 'SALES VOLUME DISTRIBUTION BY MARKET ATTRACTIVENESS'
        ws.cell(2, 1).value =  'Month'
        ws.cell(2, 2).value =  month
        ws.cell(3, 1).value =  'Figures in Ton'

        zones = final_df.ZONE.unique()
        start_col, start_row = 1, 4
        for zone in zones:
            zone_df = final_df.loc[final_df.ZONE == zone]
            ws.cell(start_row+1, start_col).value = zone
            mark_luc = zone_df.MARKET_LUCRATIVENESS.unique()
            for ix, val in enumerate(mark_luc):
                ws.cell(start_row+ix+2, start_col).value = val    

            states = zone_df.STATE.unique()
            state_row, state_col =  start_row, start_col+1
            for state in states:
                state_df = zone_df.loc[zone_df.STATE==state]
                ws.cell(state_row, state_col).value = state
                ws.cell(state_row+1, state_col).value = 'Sales Volume'
                ws.cell(state_row+1, state_col+1).value = 'Market Potential'
                ws.cell(state_row+1, state_col+2).value = 'Market Share'
                ws.cell(state_row+1, state_col+3).value = 'Sales Proportion'
                
                
                ws.cell(state_row+2, state_col).value = get_value(df=state_df, strategy='Highly Lucrative', value_to_fetch='SALES')
                ws.cell(state_row+3, state_col).value = get_value(df=state_df, strategy='Profit Driver', value_to_fetch='SALES')
                ws.cell(state_row+4, state_col).value = get_value(df=state_df, strategy='Volume Driver', value_to_fetch='SALES')
                ws.cell(state_row+5, state_col).value = get_value(df=state_df, strategy='Not Lucrative', value_to_fetch='SALES')
                ws.cell(state_row+6, state_col).value = get_value(df=state_df, strategy='Grand Total', value_to_fetch='SALES')
                
                
                
                ws.cell(state_row+2, state_col+1).value = get_value(df=state_df, strategy='Highly Lucrative', value_to_fetch='MARKET_POTENTIAL')
                ws.cell(state_row+3, state_col+1).value = get_value(df=state_df, strategy='Profit Driver', value_to_fetch='MARKET_POTENTIAL')
                ws.cell(state_row+4, state_col+1).value = get_value(df=state_df, strategy='Volume Driver', value_to_fetch='MARKET_POTENTIAL')
                ws.cell(state_row+5, state_col+1).value = get_value(df=state_df, strategy='Not Lucrative', value_to_fetch='MARKET_POTENTIAL')
                ws.cell(state_row+6, state_col+1).value = get_value(df=state_df, strategy='Grand Total', value_to_fetch='MARKET_POTENTIAL')
                
                
                ws.cell(state_row+2, state_col+2).value = get_value(df=state_df, strategy='Highly Lucrative', value_to_fetch='MARKET SHARE')
                ws.cell(state_row+3, state_col+2).value = get_value(df=state_df, strategy='Profit Driver', value_to_fetch='MARKET SHARE')
                ws.cell(state_row+4, state_col+2).value = get_value(df=state_df, strategy='Volume Driver', value_to_fetch='MARKET SHARE')
                ws.cell(state_row+5, state_col+2).value = get_value(df=state_df, strategy='Not Lucrative', value_to_fetch='MARKET SHARE')
                ws.cell(state_row+6, state_col+2).value = get_value(df=state_df, strategy='Grand Total', value_to_fetch='MARKET SHARE')
                
                
                ws.cell(state_row+2, state_col+3).value = get_value(df=state_df, strategy='Highly Lucrative', value_to_fetch='SALES_PROPORTION')
                ws.cell(state_row+3, state_col+3).value = get_value(df=state_df, strategy='Profit Driver', value_to_fetch='SALES_PROPORTION')
                ws.cell(state_row+4, state_col+3).value = get_value(df=state_df, strategy='Volume Driver', value_to_fetch='SALES_PROPORTION')
                ws.cell(state_row+5, state_col+3).value = get_value(df=state_df, strategy='Not Lucrative', value_to_fetch='SALES_PROPORTION')
                ws.cell(state_row+6, state_col+3).value = get_value(df=state_df, strategy='Grand Total', value_to_fetch='SALES_PROPORTION')
                
                
                state_col +=4
            
            ws.cell(state_row+1, state_col).value = 'Sales Volume'
            ws.cell(state_row+1, state_col+1).value = 'Market Potential'
            ws.cell(state_row+1, state_col+2).value = 'Market Share'
            ws.cell(state_row+1, state_col+3).value = 'Sales Proportion'
            zone_data = all_zones.loc[all_zones.ZONE ==zone]
            ws.cell(state_row, state_col).value = 'TOTAL'
            
            ws.cell(state_row+2, state_col).value = get_value(df=zone_data, strategy='Highly Lucrative', value_to_fetch='SALES')
            ws.cell(state_row+3, state_col).value = get_value(df=zone_data, strategy='Profit Driver', value_to_fetch='SALES')
            ws.cell(state_row+4, state_col).value = get_value(df=zone_data, strategy='Volume Driver', value_to_fetch='SALES')
            ws.cell(state_row+5, state_col).value = get_value(df=zone_data, strategy='Not Lucrative', value_to_fetch='SALES')
            ws.cell(state_row+6, state_col).value = get_value(df=zone_data, strategy='Grand Total', value_to_fetch='SALES')
            
            
            
            ws.cell(state_row+2, state_col+1).value = get_value(df=zone_data, strategy='Highly Lucrative', value_to_fetch='MARKET_POTENTIAL')
            ws.cell(state_row+3, state_col+1).value = get_value(df=zone_data, strategy='Profit Driver', value_to_fetch='MARKET_POTENTIAL')
            ws.cell(state_row+4, state_col+1).value = get_value(df=zone_data, strategy='Volume Driver', value_to_fetch='MARKET_POTENTIAL')
            ws.cell(state_row+5, state_col+1).value = get_value(df=zone_data, strategy='Not Lucrative', value_to_fetch='MARKET_POTENTIAL')
            ws.cell(state_row+6, state_col+1).value = get_value(df=zone_data, strategy='Grand Total', value_to_fetch='MARKET_POTENTIAL')
            
            
            ws.cell(state_row+2, state_col+2).value = get_value(df=zone_data, strategy='Highly Lucrative', value_to_fetch='MARKET SHARE')
            ws.cell(state_row+3, state_col+2).value = get_value(df=zone_data, strategy='Profit Driver', value_to_fetch='MARKET SHARE')
            ws.cell(state_row+4, state_col+2).value = get_value(df=zone_data, strategy='Volume Driver', value_to_fetch='MARKET SHARE')
            ws.cell(state_row+5, state_col+2).value = get_value(df=zone_data, strategy='Not Lucrative', value_to_fetch='MARKET SHARE')
            ws.cell(state_row+6, state_col+2).value = get_value(df=zone_data, strategy='Grand Total', value_to_fetch='MARKET SHARE')
            
            ws.cell(state_row+2, state_col+3).value = get_value(df=zone_data, strategy='Highly Lucrative', value_to_fetch='SALES_PROPORTION')
            ws.cell(state_row+3, state_col+3).value = get_value(df=zone_data, strategy='Profit Driver', value_to_fetch='SALES_PROPORTION')
            ws.cell(state_row+4, state_col+3).value = get_value(df=zone_data, strategy='Volume Driver', value_to_fetch='SALES_PROPORTION')
            ws.cell(state_row+5, state_col+3).value = get_value(df=zone_data, strategy='Not Lucrative', value_to_fetch='SALES_PROPORTION')
            ws.cell(state_row+6, state_col+3).value = get_value(df=zone_data, strategy='Grand Total', value_to_fetch='SALES_PROPORTION')
                
            start_row +=8
            # do for all india here
                
        ws.cell(start_row+1, start_col+1).value = 'Sales Volume'
        ws.cell(start_row+1, start_col+2).value = 'Market Potential'
        ws.cell(start_row+1, start_col+3).value = 'Market Share'
        ws.cell(start_row+1, start_col+4).value = 'Sales Proportion'
        ws.cell(start_row+1, start_col).value = 'ALL INDIA'
        ws.cell(start_row, start_col+1).value = 'ALL INDIA'
        mark_luc = all_india.MARKET_LUCRATIVENESS.unique()
        for ix, val in enumerate(mark_luc):
            ws.cell(start_row+ix+2, start_col).value = val   
            
            
        ws.cell(start_row+2, start_col+1).value = get_value(df=all_india, strategy='Highly Lucrative', value_to_fetch='SALES')
        ws.cell(start_row+3, start_col+1).value = get_value(df=all_india, strategy='Profit Driver', value_to_fetch='SALES')
        ws.cell(start_row+4, start_col+1).value = get_value(df=all_india, strategy='Volume Driver', value_to_fetch='SALES')
        ws.cell(start_row+5, start_col+1).value = get_value(df=all_india, strategy='Not Lucrative', value_to_fetch='SALES')
        ws.cell(start_row+6, start_col+1).value = get_value(df=all_india, strategy='Grand Total', value_to_fetch='SALES')



        ws.cell(start_row+2, start_col+1+1).value = get_value(df=all_india, strategy='Highly Lucrative', value_to_fetch='MARKET_POTENTIAL')
        ws.cell(start_row+3, start_col+1+1).value = get_value(df=all_india, strategy='Profit Driver', value_to_fetch='MARKET_POTENTIAL')
        ws.cell(start_row+4, start_col+1+1).value = get_value(df=all_india, strategy='Volume Driver', value_to_fetch='MARKET_POTENTIAL')
        ws.cell(start_row+5, start_col+1+1).value = get_value(df=all_india, strategy='Not Lucrative', value_to_fetch='MARKET_POTENTIAL')
        ws.cell(start_row+6, start_col+1+1).value = get_value(df=all_india, strategy='Grand Total', value_to_fetch='MARKET_POTENTIAL')

        ws.cell(start_row+2, start_col+1+2).value = get_value(df=all_india, strategy='Highly Lucrative', value_to_fetch='MARKET SHARE')
        ws.cell(start_row+3, start_col+1+2).value = get_value(df=all_india, strategy='Profit Driver', value_to_fetch='MARKET SHARE')
        ws.cell(start_row+4, start_col+1+2).value = get_value(df=all_india, strategy='Volume Driver', value_to_fetch='MARKET SHARE')
        ws.cell(start_row+5, start_col+1+2).value = get_value(df=all_india, strategy='Not Lucrative', value_to_fetch='MARKET SHARE')
        ws.cell(start_row+6, start_col+1+2).value = get_value(df=all_india, strategy='Grand Total', value_to_fetch='MARKET SHARE')

        ws.cell(start_row+2, start_col+1+3).value = get_value(df=all_india, strategy='Highly Lucrative', value_to_fetch='SALES_PROPORTION')
        ws.cell(start_row+3, start_col+1+3).value = get_value(df=all_india, strategy='Profit Driver', value_to_fetch='SALES_PROPORTION')
        ws.cell(start_row+4, start_col+1+3).value = get_value(df=all_india, strategy='Volume Driver', value_to_fetch='SALES_PROPORTION')
        ws.cell(start_row+5, start_col+1+3).value = get_value(df=all_india, strategy='Not Lucrative', value_to_fetch='SALES_PROPORTION')
        ws.cell(start_row+6, start_col+1+3).value = get_value(df=all_india, strategy='Grand Total', value_to_fetch='SALES_PROPORTION')


        wb.save(file_path)

    def get_run_dtl(cnxn,month):
        sql = f'''
            select * from etl_zone."MARKET_MAPPING_RUN" mmr 
            where "PLAN_MONTH" = '{month}'
            order by "RUN_ID" desc 
            limit 1
        '''

        run_dtl = pd.read_sql(sql,cnxn)

        return run_dtl

    def get_data(cnxn,run_id):

        sql = f'''
            select * from etl_zone."MARKET_MAPPING_DISTRICT_CLASSIFICATION" mmdc 
            where "RUN_ID" = {run_id}
        '''

        df = pd.read_sql(sql,cnxn)

        return df

    # if __name__=='__main__':
    #     cnxn = connect_db()
        
    #     month = '2023-03-01'

    #     try:
    #         run_dtl = get_run_dtl(cnxn,month)
    #         print(run_dtl)
    #         df = get_data(cnxn,run_dtl.loc[0,'RUN_ID'])
    #         # df = get_data(cnxn,65)

    #         get_file(df, month)
            
    #     except Exception as e:
    #         print(e)
    #     finally:
    #         cnxn.close()
