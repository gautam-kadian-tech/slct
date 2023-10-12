"""Handling agent views module."""
import calendar
import json
from datetime import date
from datetime import date as datetime_date
from datetime import datetime
from datetime import datetime as dt
from datetime import timedelta
from datetime import timedelta as td
from io import BytesIO

import pandas as pd
import psycopg2
from dateutil.relativedelta import relativedelta
from django.conf import settings
from django.db import transaction
from django.db.models import (
    Avg,
    Case,
    CharField,
    Count,
    DecimalField,
    DurationField,
    ExpressionWrapper,
    F,
    Q,
    Sum,
    Value,
    When,
    functions,
)
from django.db.models.functions import ExtractMonth, ExtractWeek, Substr
from django.http import HttpResponse
from django_filters.rest_framework import DjangoFilterBackend
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from rest_framework.filters import OrderingFilter, SearchFilter
from rest_framework.generics import GenericAPIView, ListAPIView
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.viewsets import ModelViewSet

from accounts.models import TgtRlsRoleData
from accounts.user_role_choices import UserRoleChoice
from analytical_data.custom_mixins import RequestsCountMixin
from analytical_data.custom_permissions import (
    IsHandlingAgent,
    IsLogisticsHead,
    IsRRCCement,
)
from analytical_data.enum_classes import ApprovalStatusChoices
from analytical_data.filters import (
    EpodDataFilter,
    FreightChangeInitiationFilter,
    GodownPerformanceFilter,
    HandlingAgentDashboardFilter,
    NewFreightInitiationFilter,
    SidingConstraintsFilterset,
    TgtDayWiseLiftingFilter,
    TgtDepoInventoryStkFilter,
    TgtMrnDataFilter,
    TgtPlantDispatchDataFilter,
    TgtPlantDispatchDataFilternew,
    TgtPlantSiloCapacityFilter,
    TgtRakeLoadingDetailsFilter,
    TgtRakeLoadingFilter,
    TgtRakeUnloadingDetailsFilter,
    TgtSlhOrderPendencyFilter,
    TgtSlhServiceLevelDepoFilter,
)
from analytical_data.filters.handling_agent_filters import (
    RailExpensesDetailsFilterset,
    RailExpensesDetailsWfFilterset,
)
from analytical_data.models import (
    ApprovalThreshold,
    CrwcChargesMaster,
    DemurrageSlabs,
    DepoWiseFreightMaster,
    EpodData,
    FreightChangeInitiation,
    FreightDiscoveryProfiles,
    GdWharfageOutput,
    HandlingMasters,
    HourlyLiftingEfficiency,
    HourlyLiftingEfficiencyMaster,
    LiftingPattern,
    NewFreightInitiation,
    RailExpensesDetails,
    RailExpensesDetailsWarfage,
    RakeHandlingMaster,
    ReasonsForDemurrageWharfage,
    ReasonsForFreightChange,
    SidingConstraints,
    SidingWiseLiasioningAgent,
    TgtDayWiseLifting,
    TgtDepoDispatchData,
    TgtDepoInventoryStk,
    TgtMrnData,
    TgtPlantDepoMaster,
    TgtPlantDispatchData,
    TgtPlantSiloCapacity,
    TgtRakeCharges,
    TgtRakeDisposals,
    TgtRakeLoading,
    TgtRakeLoadingDetails,
    TgtRakeUnloadingDetails,
    TgtSlhOrderPendency,
    TgtSlhServiceLevelDepo,
    TOebsFndLookupValues,
    WaiverCommissionMaster,
    WharfageSlabs,
)
from analytical_data.models.handling_agent_models import RrcIsoStockTransfer
from analytical_data.models.monthly_scheduling_models import (
    LpModelDfFnl,
    LpModelRun,
)
from analytical_data.serializers import (
    CostsMasterDetailSerializer,
    CrwcChargesMasterSerializer,
    DemurrageSlabsSerializer,
    EpodDataSerializer,
    FreightChangeInitiationSerializer,
    GdVsWharfageInputSerializer,
    GdVsWharfageModelRunOutputByRunidSerializer,
    GdVsWhargfageOutputSerializer,
    HandlingMastersSerializer,
    HourlyLiftingEfficiencyMasterSerializer,
    HourlyLiftingEfficiencySerializer,
    LiftingPatternSerializer,
    NewFreightInitiationSerializer,
    RailExpensesDetailsPostSerializer,
    RailExpensesDetailsSerializer,
    RailExpensesDetailsWfPostSerializer,
    RailExpensesDetailsWfSerializer,
    ReasonsForDemurrageWharfageSerializer,
    RoadRakeCoordinatorSerializer,
    SidingConstraintsSerializer,
    TgtDayWiseLiftingSerializer,
    TgtDepoInventoryStkSerializer,
    TgtMrnDataSerializer,
    TgtPlantDispatchDataSerializer,
    TgtPlantSiloCapacitySerializer,
    TgtRakeChargesSerializer,
    TgtRakeDisposalsSerializer,
    TgtRakeLoadingDetailsSerializer,
    TgtRakeLoadingDetailsSumSerializer,
    TgtRakeLoadingGetSerializer,
    TgtRakeLoadingReadOnlySerializer,
    TgtRakeLoadingSerializer,
    TgtRakeUnloadingDetailsSerializer,
    TgtSlhOrderPendencySerializer,
    TgtSlhServiceLevelDepoSerializer,
    WaiverCommissionMasterSerializer,
    WharfageSlabsSerializer,
)
from analytical_data.utils import CustomPagination, Responses
from analytical_data.view_helpers import GdVsWharfageHelper
from analytical_data.view_helpers.connection import connect_db
from analytical_data.views import DownloadUploadViewSet


class HandingAgentDashboardDropdown(GenericAPIView):
    queryset = TgtPlantDepoMaster.objects.all()
    filter_backends = (SearchFilter,)
    search_fields = ("party_name",)

    def get(self, request, *args, **kwargs):
        depo = (
            self.filter_queryset(self.get_queryset())
            .filter(type="DEPO")
            .values_list("party_name")
            .annotate(Count("party_name"), depo=Substr("party_name", pos=1, length=3))
            .values_list("party_name", flat=True)
        )

        return Responses.success_response("Handling agent dropdown data.", data=depo)


class HandlingAgentDashboard(GenericAPIView):
    queryset = TgtDepoDispatchData.objects.all()
    filter_backends = (DjangoFilterBackend,)
    filterset_class = HandlingAgentDashboardFilter

    def get(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())
        if not request.query_params.get("tax_invoice_date"):
            return Responses.error_response(
                "tax_invoice_date is a required query_param filter"
            )
        tax_invoice_date = datetime.strptime(
            request.query_params.get("tax_invoice_date"), "%Y-%m-%d"
        )

        month_to_date_query = Q(
            tax_invoice_date__date__range=[
                tax_invoice_date.replace(
                    day=1, hour=0, minute=0, second=0, microsecond=0
                ),
                tax_invoice_date,
            ]
        )

        godown_dispatch_qty = (
            queryset.filter(tax_invoice_date__date=tax_invoice_date)
            .aggregate(Sum("shipped_qty"))
            .get("shipped_qty__sum")
        )
        godown_dispatch_qty_mtd = (
            queryset.filter(month_to_date_query)
            .aggregate(Sum("shipped_qty"))
            .get("shipped_qty__sum")
        )

        crossing_qty = (
            queryset.filter(
                tax_invoice_date__date=tax_invoice_date,
                sale_type="TP",
            )
            .aggregate(Sum("shipped_qty"))
            .get("shipped_qty__sum")
        )
        crossing_qty_mtd = (
            queryset.filter(
                month_to_date_query,
                sale_type="TP",
            )
            .aggregate(Sum("shipped_qty"))
            .get("shipped_qty__sum")
        )

        total_dispatch_qty_road = (
            queryset.filter(
                ~Q(sale_type__in=("GD-RK", "RK")),
                tax_invoice_date__date=tax_invoice_date,
                mode_of_transport="ROAD",
            )
            .aggregate(Sum("shipped_qty"))
            .get("shipped_qty__sum")
        )
        total_dispatch_qty_mtd_road = (
            queryset.filter(
                month_to_date_query,
                ~Q(sale_type__in=("GD-RK", "RK")),
                mode_of_transport="ROAD",
            )
            .aggregate(Sum("shipped_qty"))
            .get("shipped_qty__sum")
        )

        total_dispatch_qty_rail = (
            queryset.filter(
                tax_invoice_date__date=tax_invoice_date,
                sale_type__in=("GD-RK", "RK"),
                mode_of_transport="ROAD",
            )
            .aggregate(Sum("shipped_qty"))
            .get("shipped_qty__sum")
        )
        total_dispatch_qty_mtd_rail = (
            queryset.filter(
                month_to_date_query,
                sale_type__in=("GD-RK", "RK"),
                mode_of_transport="ROAD",
            )
            .aggregate(Sum("shipped_qty"))
            .get("shipped_qty__sum")
        )

        diversion_qty = (
            queryset.filter(tax_invoice_date__date=tax_invoice_date, sale_type="RD")
            .aggregate(Sum("shipped_qty"))
            .get("shipped_qty__sum")
        )
        diversion_qty_mtd = (
            queryset.filter(month_to_date_query, sale_type="RD")
            .aggregate(Sum("shipped_qty"))
            .get("shipped_qty__sum")
        )

        rh_fired_qty = (
            queryset.filter(tax_invoice_date__date=tax_invoice_date, sale_type="RK")
            .aggregate(Sum("shipped_qty"))
            .get("shipped_qty__sum")
        )
        rh_fired_qty_mtd = (
            queryset.filter(month_to_date_query, sale_type="RK")
            .aggregate(Sum("shipped_qty"))
            .get("shipped_qty__sum")
        )
        qty_inward_at_depot_mtd = (
            queryset.filter(month_to_date_query, sale_type__in=("GD-RK", "RK"))
            .aggregate(Sum("shipped_qty"))
            .get("shipped_qty__sum")
        )

        data = {
            "godown_dispatch_qty": godown_dispatch_qty,
            "godown_dispatch_qty_mtd": godown_dispatch_qty_mtd,
            "crossing_qty": crossing_qty,
            "crossing_qty_mtd": crossing_qty_mtd,
            "diversion_qty": diversion_qty,
            "diversion_qty_mtd": diversion_qty_mtd,
            "rh_fired_qty": rh_fired_qty,
            "rh_fired_qty_mtd": rh_fired_qty_mtd,
            "total_dispatch_qty_road": total_dispatch_qty_road,
            "total_dispatch_qty_mtd_road": total_dispatch_qty_mtd_road,
            "total_dispatch_qty_rail": total_dispatch_qty_rail,
            "total_dispatch_qty_mtd_rail": total_dispatch_qty_mtd_rail,
        }
        return Responses.success_response("Handling agent dashboard data.", data=data)


class TGTPlantSiloCapacityViewSet(ModelViewSet):
    """TGT Plant Silo Capacity CRUDs view set class."""

    queryset = TgtPlantSiloCapacity.objects.all()
    serializer_class = TgtPlantSiloCapacitySerializer
    filter_backends = (DjangoFilterBackend,)
    filterset_class = TgtPlantSiloCapacityFilter
    search_fields = ("code", "product")
    pagination_class = CustomPagination
    lookup_field = "id"


class TgtPlantSiloCapacityDropdown(GenericAPIView):
    queryset = TgtPlantSiloCapacity.objects.all()
    filter_backends = (DjangoFilterBackend,)
    filterset_class = TgtPlantSiloCapacityFilter

    def __get_silo_capacity_query(self, query_string):
        return (
            self.filter_queryset(self.get_queryset())
            .values_list(query_string, flat=True)
            .annotate(Count(query_string))
            .order_by(query_string)
        )

    def get(self, request, *args, **kwargs):
        data = {
            "plant": self.__get_silo_capacity_query("code"),
            "product": self.__get_silo_capacity_query("product"),
        }
        return Responses.success_response("tgt silo capacity dropdown data.", data=data)


class GodownPerformance(GenericAPIView):
    queryset = TgtDepoDispatchData.objects.all()
    filter_backends = (DjangoFilterBackend,)
    filterset_class = GodownPerformanceFilter

    def get(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())
        depo = request.GET.get("depo")
        tax_invoice_date = request.GET.get("tax_invoice_date")
        tax_invoice_date = datetime.strptime(tax_invoice_date, "%Y-%m-%d")
        month_of_tax_invoice_date = tax_invoice_date.month
        run_id_dict = (
            LpModelRun.objects.filter(
                plan_date__month=month_of_tax_invoice_date, run_status=1
            )
            .values("run_id")
            .first()
        )
        if run_id_dict is not None:
            run_id = run_id_dict["run_id"]
        else:
            run_id = None

        godown_dispatch_qty = LpModelDfFnl.objects.filter(
            warehouse=depo, primary_secondary_route="SECONDARY"
        ).aggregate(
            shipped_qty__sum=functions.Coalesce(
                Sum("qty"), 0.0, output_field=DecimalField()
            )
        )[
            "shipped_qty__sum"
        ]
        # godown_dispatch_qty = queryset.aggregate(
        #     shipped_qty__sum=functions.Coalesce(
        #         Sum("shipped_qty"), 0.0, output_field=DecimalField()
        #     )
        # )["shipped_qty__sum"]
        target_qty = LpModelDfFnl.objects.filter(
            warehouse=depo, run_id=run_id, primary_secondary_route="SECONDARY"
        ).aggregate(
            shipped_qty__sum=functions.Coalesce(
                Sum("qty"), 0.0, output_field=DecimalField()
            )
        )[
            "shipped_qty__sum"
        ]
        dispatch_data_by_sale_type = (
            queryset.values("sale_type")
            .annotate(
                Count("sale_type"),
                Sum("shipped_qty"),
            )
            .values(
                "sale_type",
                "shipped_qty__sum",
            )
        )
        rake_sale_type_count = queryset.filter(sale_type__in=("RK", "GD-RK")).aggregate(
            Sum("shipped_qty"),
        )
        road_sale_type_count = queryset.exclude(
            sale_type__in=("RK", "GD-RK")
        ).aggregate(
            Sum("shipped_qty"),
        )

        data = {
            "godown_dispatch_qty": round(godown_dispatch_qty, 2),
            "target_qty": target_qty,
            "dispatch_data_by_sale_type": dispatch_data_by_sale_type,
            "rake_qty_sum": rake_sale_type_count["shipped_qty__sum"],
            "road_qty_sum": road_sale_type_count["shipped_qty__sum"],
        }
        return Responses.success_response("Handling agent dashboard data.", data=data)


class GodownTat(GenericAPIView):
    queryset = TgtSlhServiceLevelDepo.objects.all()
    filter_backends = (DjangoFilterBackend,)
    filterset_class = TgtSlhServiceLevelDepoFilter

    def get(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())
        dispatch_tat = queryset.aggregate(
            total_time=functions.Coalesce(
                Avg(F("tax_invoice_date") - F("oe_creation_date")),
                timedelta(hours=0),
                output_field=DurationField(),
            )
        ).get("total_time")
        plan_dispatch_tat = 1
        service_sla = queryset.aggregate(
            total_time=functions.Coalesce(
                Avg(F("epod1_time") - F("oe_creation_date")),
                timedelta(0),
                output_field=DurationField(),
            )
        ).get("total_time")
        plan_service_sla = queryset.aggregate(
            plan_service_sla=functions.Coalesce(
                Avg(Value(plan_dispatch_tat) + ((F("distance") / 25))),
                0.0,
                output_field=DecimalField(),
            )
        ).get("plan_service_sla")

        data = {
            "dispatch_tat": round(
                getattr(dispatch_tat, "days", 0) * 24
                + getattr(dispatch_tat, "seconds", 0) / 3600,
                1,
            ),
            "plan_dispatch_tat": plan_dispatch_tat,
            "service_sla": round(
                getattr(service_sla, "days", 0) * 24
                + getattr(service_sla, "seconds", 0) / 3600,
                1,
            ),
            "plan_service_sla": round(plan_service_sla, 1),
        }
        return Responses.success_response("Handling agent dashboard data.", data=data)


# class GodownPerformanceDispatchTatAndServiceSla(GenericAPIView):

#     queryset = TgtDepoInventoryStk.objects.all()
#     filter_backends = (DjangoFilterBackend,)
#     filterset_fields = ("plant_depo", "di_date")

#     def get(self, request, *args, **kwargs):
#         queryset = self.filter_queryset(self.get_queryset())
#         inventory_data = queryset.aggregate(Sum("shipped_qty")).get(
#             "shipped_qty__sum"
#         )
#         data = {
#             # "dispatch_data_by_sale_type": dispatch_data_by_sale_type,
#         }
#         return Responses.success_response("Handling agent dashboard data.", data=data)


class RoadRakeCoordinatorAPIView(ModelViewSet):
    queryset = FreightDiscoveryProfiles.objects.all()
    serializer_class = RoadRakeCoordinatorSerializer
    file_name = "freight_discovery"
    lookup_field = "id"

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        request.session["freight_discovery"] = serializer.validated_data
        return Responses.success_response(
            "Road rake coordinator analysis screen data.",
            data=serializer.validated_data,
        )

    def update(self, request, *args, **kwargs):
        serializer = self.get_serializer(self.get_object(), data=request.data)
        serializer.is_valid(raise_exception=True)
        self.perform_update(serializer)
        request.session["freight_discovery"] = serializer.data
        return Response(serializer.data)

    def download(self, request, *args, **kwargs):
        data = self.get_serializer(self.get_object()).data
        bio = BytesIO()
        workbook = Workbook()
        ws = workbook["Sheet"]

        green = PatternFill(start_color="8AFF8A", end_color="8AFF8A", fill_type="solid")
        dark_green = PatternFill(
            start_color="59981A", end_color="59981A", fill_type="solid"
        )
        bold_font = Font(bold=True)

        ws.cell(2, 2).value = "EY FREIDISC - Freight Discovery Model"
        ws.cell(2, 2).font = bold_font
        ws.cell(4, 2).value = "Overview"
        ws.cell(4, 2).font = bold_font

        ws.cell(6, 2).value = "Origin"
        ws.cell(6, 2).font = bold_font
        row, column = 7, 2

        column_len = 0

        for key, value in data.get("origin").items():
            if len(key) > column_len:
                column_len = len(key)
            ws.cell(row, column).value = key
            ws.cell(row, column + 1).value = value
            row += 1
        row += 1

        ws.cell(row + 1, column).value = "Destination"
        ws.cell(row + 1, column).font = bold_font
        row += 2
        for key, value in data.get("destination").items():
            if len(key) > column_len:
                column_len = len(key)
            ws.cell(row, column).value = key
            ws.cell(row, column + 1).value = value
            row += 1
        row += 1

        ws.cell(row + 1, column).value = "Cost Profile"
        ws.cell(row + 1, column).font = bold_font
        row += 2
        for key, value in data.get("cost_profile").items():
            if len(key) > column_len:
                column_len = len(key)
            ws.cell(row, column).value = key
            ws.cell(row, column + 1).value = value
            ws.cell(row, column).fill = green
            ws.cell(row, column + 1).fill = green
            row += 1
        ws.cell(row - 1, column).fill = dark_green
        ws.cell(row - 1, column + 1).fill = dark_green
        row += 1

        ws.cell(row + 1, column).value = "Primary Assumptions"
        ws.cell(row + 1, column).font = bold_font
        row += 2
        for key, value in data.get("primary_assumptions").items():
            if len(key) > column_len:
                column_len = len(key)
            ws.cell(row, column).value = key
            ws.cell(row, column + 1).value = value
            row += 1
        row += 1

        ws.column_dimensions[get_column_letter(column)].width = column_len
        column_len = 0

        ws.cell(4, 6).value = "Financing Assumptions"
        ws.cell(4, 6).font = bold_font
        row, column = 6, 6

        ws.cell(row, column).value = "Purchase Inputs"
        ws.cell(row, column).font = bold_font
        row += 1
        for key, value in data.get("purchase_inputs").items():
            if len(key) > column_len:
                column_len = len(key)
            ws.cell(row, column).value = key
            ws.cell(row, column + 1).value = value
            row += 1
        row += 2

        ws.cell(row, column).value = "Capex Computations"
        ws.cell(row, column).font = bold_font
        row += 1
        for key, value in data.get("capex_computations").items():
            if len(key) > column_len:
                column_len = len(key)
            ws.cell(row, column).value = key
            ws.cell(row, column + 1).value = value
            row += 1
        row += 2

        ws.cell(row, column).value = "Profitability Setting"
        ws.cell(row, column).font = bold_font
        row += 1
        for key, value in data.get("profitability_settings").items():
            if len(key) > column_len:
                column_len = len(key)
            ws.cell(row, column).value = key
            ws.cell(row, column + 1).value = value
            row += 1
        row += 2

        ws.cell(row, column).value = "Financial Feasibility Check"
        ws.cell(row, column).font = bold_font
        row += 1
        for key, value in data.get("financial_feasibility_check").items():
            if len(key) > column_len:
                column_len = len(key)
            ws.cell(row, column).value = key
            ws.cell(row, column + 1).value = value
            row += 1
        row += 1

        ws.column_dimensions[get_column_letter(column)].width = column_len
        column_len = 0

        ws.cell(4, 9).value = "Direct Cost Assumptions"
        ws.cell(4, 9).font = bold_font
        row, column = 6, 9

        ws.cell(row, column).value = "Tyre Charges"
        ws.cell(row, column).font = bold_font
        row += 1
        for key, value in data.get("tyre_charges").items():
            if len(key) > column_len:
                column_len = len(key)
            ws.cell(row, column).value = key
            ws.cell(row, column + 1).value = value
            row += 1
        row += 2

        ws.cell(row, column).value = "Fuel Charges"
        ws.cell(row, column).font = bold_font
        row += 1
        for key, value in data.get("fuel_charges").items():
            if len(key) > column_len:
                column_len = len(key)
            ws.cell(row, column).value = key
            ws.cell(row, column + 1).value = value
            row += 1
        row += 2

        ws.cell(row, column).value = "Direct Cost Computation"
        ws.cell(row, column).font = bold_font
        row += 1
        for key, value in data.get("direct_cost_computations").items():
            if len(key) > column_len:
                column_len = len(key)
            ws.cell(row, column).value = key
            ws.cell(row, column + 1).value = value
            row += 1
        row += 2

        ws.column_dimensions[get_column_letter(column)].width = column_len
        column_len = 0

        ws.cell(4, 12).value = "Direct Cost Assumptions"
        ws.cell(4, 12).font = bold_font
        row, column = 6, 12

        ws.cell(row, column).value = "Other Variable Charges"
        ws.cell(row, column).font = bold_font
        row += 1
        for key, value in data.get("other_variable_charges").items():
            if len(key) > column_len:
                column_len = len(key)
            ws.cell(row, column).value = key
            ws.cell(row, column + 1).value = value
            row += 1
        row += 2

        ws.cell(row, column).value = "Semi Variable Cost Computation"
        ws.cell(row, column).font = bold_font
        row += 1
        for key, value in data.get("semi_variable_cost_computations").items():
            if len(key) > column_len:
                column_len = len(key)
            ws.cell(row, column).value = key
            ws.cell(row, column + 1).value = value
            row += 1
        row += 2

        ws.column_dimensions[get_column_letter(column)].width = column_len
        column_len = 0

        ws.cell(4, 15).value = "Fixed Cost Computation"
        ws.cell(4, 15).font = bold_font
        row, column = 6, 15

        ws.cell(row, column).value = "Annual Charges"
        ws.cell(row, column).font = bold_font
        row += 1
        for key, value in data.get("annual_charges").items():
            if len(key) > column_len:
                column_len = len(key)
            ws.cell(row, column).value = key
            ws.cell(row, column + 1).value = value
            row += 1
        row += 2

        ws.cell(row, column).value = "Monthly Charges"
        ws.cell(row, column).font = bold_font
        row += 1
        for key, value in data.get("monthly_charges").items():
            if len(key) > column_len:
                column_len = len(key)
            ws.cell(row, column).value = key
            ws.cell(row, column + 1).value = value
            row += 1
        row += 2

        ws.cell(row, column).value = "Fixed Cost Computation"
        ws.cell(row, column).font = bold_font
        row += 1
        for key, value in data.get("fixed_cost_computations").items():
            if len(key) > column_len:
                column_len = len(key)
            ws.cell(row, column).value = key
            ws.cell(row, column + 1).value = value
            row += 1
        row += 2

        ws.column_dimensions[get_column_letter(column)].width = column_len
        column_len = 0

        workbook.save(bio)
        bio.seek(0)
        content_type = (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response = HttpResponse(bio.getvalue(), content_type=content_type)
        response["Content-Disposition"] = f"attachment; filename={self.file_name}.xlsx"
        return response


class TgtSlhOrderPendencyViewSet(ModelViewSet):
    queryset = TgtSlhOrderPendency.objects.all()
    serializer_class = TgtSlhOrderPendencySerializer
    filter_backends = (DjangoFilterBackend,)
    filterset_class = TgtSlhOrderPendencyFilter
    pagination_class = CustomPagination

    def filter_queryset(self, queryset):
        queryset = super().filter_queryset(queryset)
        status = self.request.query_params.get("status")

        if status == "pending":
            return queryset.filter(order_status="PENDING ORDER")

        if status == "ready_for_delivery":
            return queryset.filter(~Q(order_status="PENDING ORDER"))

        return queryset


class TgtMrnDataViewSet(ModelViewSet):
    queryset = TgtMrnData.objects.all()
    serializer_class = TgtMrnDataSerializer
    pagination_class = CustomPagination
    filter_backends = (DjangoFilterBackend,)
    filterset_class = TgtMrnDataFilter

    def filter_queryset(self, queryset):
        queryset = super().filter_queryset(queryset)
        status = self.request.query_params.get("status")
        order_quantity = self.request.query_params.get("sorting_on", "id")

        if status == "Received":
            return queryset.filter(receipt_number__isnull=False).order_by(
                order_quantity
            )

        if status == "Ship":
            return queryset.filter(
                # delivery_id__isnull=False,
                # excise_invoice_date__isnull=True,
                actual_departure_date__isnull=False,
                receipt_date__isnull=True,
            ).order_by(order_quantity)

        if status == "Pending":
            return queryset.filter(
                receipt_number__isnull=True,
                excise_invoice_date__isnull=False,
            ).order_by(order_quantity)
        return queryset

    def list(self, request, *args, **kwargs):
        response = super().list(request, *args, **kwargs)

        response.data["data"]["_meta"]["total_orders"] = len(
            self.filter_queryset(self.get_queryset())
            .order_by()
            .values_list("order_number", flat=True)
            .annotate(Count("order_number"))
        )
        response.data["data"]["_meta"]["total_quantity"] = self.filter_queryset(
            self.queryset
        ).aggregate(Sum("ordered_quantity"))["ordered_quantity__sum"]
        return response


class TgtSlhServiceLevelDepoAPIView(ListAPIView):
    queryset = TgtSlhServiceLevelDepo.objects.all()
    serializer_class = TgtSlhServiceLevelDepoSerializer
    filter_backends = (DjangoFilterBackend, OrderingFilter)
    filterset_class = TgtSlhServiceLevelDepoFilter
    ordering_fields = ("dispatched_qty", "oe_creation_dt")
    pagination_class = CustomPagination


class AllproductsDropdown(GenericAPIView):
    def get(self, request):
        model = request.query_params.get("model")
        if model == "inventory":
            tgt_depo_inventory = TgtDepoInventoryStk.objects.distinct(
                "item"
            ).values_list("item", flat=True)
            return Responses.success_response(
                "tgt depo inventory stk", data=tgt_depo_inventory
            )
        elif model == "tgt_mrn":
            tgt_mrn_data = TgtMrnData.objects.distinct("ordered_item").values_list(
                "ordered_item", flat=True
            )
            return Responses.success_response("tgt mrn data", data=tgt_mrn_data)
        elif model == "depot_dispatch":
            tgt_depo_dispatch = TgtDepoDispatchData.objects.distinct(
                "product"
            ).values_list("product", flat=True)
            return Responses.success_response(
                "depo dispatch data", data=tgt_depo_dispatch
            )
        else:
            tgt_depo_inventory = TgtDepoInventoryStk.objects.distinct(
                "item"
            ).values_list("item", flat=True)
            tgt_mrn_data = TgtMrnData.objects.distinct("ordered_item").values_list(
                "ordered_item", flat=True
            )
            tgt_depo_dispatch = TgtDepoDispatchData.objects.distinct(
                "product"
            ).values_list("product", flat=True)
            final_list = set(
                list(tgt_depo_inventory) + list(tgt_mrn_data) + list(tgt_depo_dispatch)
            )
            return Responses.success_response(
                "all product data dropdown",
                data={"all_products_list": list(final_list)},
            )


class WaterFallChartInventoryDropdownViewSet(ModelViewSet):
    def get(self, request):
        depo = request.query_params.get("depo")
        date = datetime.strptime(request.query_params.get("date"), "%Y-%m-%d").date()
        inventory_stk_queryset = (
            TgtDepoInventoryStk.objects.filter(whse_code=depo, trans_date=date)
            .distinct("item")
            .values_list("item", flat=True)
        )
        stock_received_todays = (
            TgtMrnData.objects.filter(
                customer__startswith=depo, receipt_date__date=date
            )
            .distinct("ordered_item")
            .values_list("ordered_item", flat=True)
        )
        stock_dispatch_today = (
            TgtDepoDispatchData.objects.filter(
                excise_invoice_no__excise_invoice_no__startswith=depo,
                tax_invoice_date__date=date,
            )
            .distinct("product")
            .values_list("product", flat=True)
        )
        product_list = (
            list(inventory_stk_queryset)
            + list(stock_received_todays)
            + list(stock_dispatch_today)
        )
        final_product_list = list(set(product_list))

        return Responses.success_response(
            "WaterFall Chart Inventory Product Dropdown", data=final_product_list
        )


class WaterfallChartInventoryViewSet(GenericAPIView):
    def get(self, request):
        plant = request.query_params.get("plant")
        item_name = request.query_params.get("product")
        depo = request.query_params.get("depo")
        date = datetime.strptime(request.query_params.get("date"), "%Y-%m-%d").date()
        if item_name:
            inventory_stk_queryset = TgtDepoInventoryStk.objects.filter(
                whse_code=depo, trans_date=date, item=item_name
            ).aggregate(Sum("opening_stock"))
            if inventory_stk_queryset["opening_stock__sum"] is None:
                opening_stock_today = 0
            else:
                opening_stock_today = inventory_stk_queryset["opening_stock__sum"]
            stock_received_todays = TgtMrnData.objects.filter(
                customer__startswith=depo,
                receipt_date__date=date,
                ordered_item=item_name,
            ).aggregate(Sum("quantity_shipped"))
            if stock_received_todays["quantity_shipped__sum"] is None:
                stock_received_today = 0
            else:
                (stock_received_today) = stock_received_todays["quantity_shipped__sum"]
            stock_dispatch_today = TgtDepoDispatchData.objects.filter(
                excise_invoice_no__excise_invoice_no__startswith=depo,
                tax_invoice_date__date=date,
                product=item_name,
            ).aggregate(Sum("ordered_quantity"))
            if stock_dispatch_today["ordered_quantity__sum"] is None:
                (stock_dispatch_today) = 0
            else:
                (stock_dispatch_today) = (
                    stock_dispatch_today["ordered_quantity__sum"] * -1
                )

            stocks_transit = TgtMrnData.objects.filter(
                customer__startswith=depo,
                actual_departure_date__isnull=False,
                receipt_date__isnull=True,
                ordered_item=item_name,
            ).aggregate(Sum("ordered_quantity"))

            if stocks_transit["ordered_quantity__sum"] is None:
                (stock_transit) = 0
            else:
                (stock_transit) = stocks_transit["ordered_quantity__sum"]
            data_dict = {
                "opening_stock_today": opening_stock_today,
                "stock_received_today": stock_received_today,
                "stock_dispatch_today": stock_dispatch_today,
                "stock_in_transit": stock_transit,
                "Stock_in_hand": round(
                    float(opening_stock_today)
                    + float(stock_received_today)
                    + float(stock_dispatch_today)
                    + float(stock_transit),
                    2,
                ),
            }
            return Responses.success_response(
                "Waterfall chart inventory data fetched successfully", data=data_dict
            )
        else:
            inventory_stk_queryset = TgtDepoInventoryStk.objects.filter(
                whse_code=depo,
                trans_date=date,
            ).aggregate(Sum("opening_stock"))
            if inventory_stk_queryset["opening_stock__sum"] is None:
                opening_stock_today = 0
            else:
                opening_stock_today = inventory_stk_queryset["opening_stock__sum"]
            stock_received_todays = (
                TgtMrnData.objects.filter(
                    customer__startswith=depo, receipt_date__date=date
                )
                .values("delivery_id")
                .distinct()
                .aggregate(Sum("quantity_shipped"))
            )
            if stock_received_todays["quantity_shipped__sum"] is None:
                stock_received_today = 0
            else:
                (stock_received_today) = stock_received_todays["quantity_shipped__sum"]
            stock_dispatch_today = TgtDepoDispatchData.objects.filter(
                excise_invoice_no__excise_invoice_no__startswith=depo,
                tax_invoice_date__date=date,
            ).aggregate(Sum("ordered_quantity"))
            if stock_dispatch_today["ordered_quantity__sum"] is None:
                (stock_dispatch_today) = 0
            else:
                (stock_dispatch_today) = (
                    stock_dispatch_today["ordered_quantity__sum"] * -1
                )

            stocks_transit = (
                TgtMrnData.objects.filter(
                    customer__startswith=depo,
                    actual_departure_date__isnull=False,
                    receipt_date__isnull=True,
                )
                .values("delivery_id")
                .distinct()
                .aggregate(Sum("ordered_quantity"))
            )

            if stocks_transit["ordered_quantity__sum"] is None:
                (stock_transit) = 0
            else:
                (stock_transit) = stocks_transit["ordered_quantity__sum"]
            data_dict = {
                "opening_stock_today": opening_stock_today,
                "stock_received_today": stock_received_today,
                "stock_dispatch_today": stock_dispatch_today,
                "stock_in_transit": stock_transit,
                "Stock_in_hand": round(
                    float(opening_stock_today)
                    + float(stock_received_today)
                    + float(stock_dispatch_today)
                    + float(stock_transit),
                    2,
                ),
            }

            return Responses.success_response(
                "Waterfall chart inventory data fetched successfully", data=data_dict
            )


# class ProductWiseInventory(ModelViewSet):
#     queryset = TgtDepoInventoryStk.objects.all()
#     serializer_class = TgtDepoInventoryStkSerializer
#     filter_backends = (DjangoFilterBackend,)
#     filterset_class = TgtDepoInventoryStkFilter
#     pagination_class = CustomPagination

#     def get(self, request, *args, **kwargs):
#         selected_date = self.request.query_params.get("date")
#         selected_date = datetime.strptime(selected_date, "%Y-%m-%d").date()
#         today = selected_date.replace(day=1)
#         month_count = 0

#         product_wise_inventory_data = []
#         dates = []

#         product_wise_inv_data_today = (
#             self.filter_queryset(self.get_queryset().filter(trans_date=selected_date))
#             .values("trans_date", "location")
#             .annotate(Count("location"), Sum("opening_stock"))
#         )
#         product_wise_inventory_data.append(product_wise_inv_data_today)
#         dates.append(selected_date)

#         while month_count < 6:
#             day = today - relativedelta(months=month_count)
#             dates.append(day)
#             month_count += 1
#             product_wise_inv_data = (
#                 self.filter_queryset(self.get_queryset().filter(trans_date=day))
#                 .values("trans_date", "location")
#                 .annotate(Count("location"), Sum("opening_stock"))
#             )
#             product_wise_inventory_data.append(product_wise_inv_data)

#         data_dict = {
#             "product_wise_inventory_data": product_wise_inventory_data,
#             "dates": dates,
#         }
#         return Responses.success_response(
#             "product wise inventory data ", data=(data_dict)
#         )


class MonthlyInventory(ModelViewSet):
    queryset = TgtDepoInventoryStk.objects.all()
    serializer_class = TgtDepoInventoryStkSerializer
    filter_backends = (DjangoFilterBackend,)
    filterset_class = TgtDepoInventoryStkFilter
    pagination_class = CustomPagination

    def get(self, request, *args, **kwargs):
        selected_date = self.request.query_params.get("date")
        selected_date = datetime.strptime(selected_date, "%Y-%m-%d").date()
        today = selected_date.replace(day=1)
        month_count = 0
        monthly_inventory_data = []
        dates = []
        product_wise_inv_data_today = (
            self.filter_queryset(self.get_queryset().filter(trans_date=selected_date))
            .values("trans_date", "item")
            .annotate(Count("item"), Sum("opening_stock"))
        )

        monthly_inventory_data.append(product_wise_inv_data_today)
        dates.append(selected_date)

        while month_count < 6:
            day = today - relativedelta(months=month_count)
            dates.append(day)
            month_count += 1
            product_wise_inv_data = (
                self.filter_queryset(self.get_queryset().filter(trans_date=day))
                .values("trans_date", "item")
                .annotate(Count("item"), Sum("opening_stock"))
            )
            monthly_inventory_data.append(product_wise_inv_data)

        data_dict = {
            "monthly_inventory_data": monthly_inventory_data,
            "dates": dates,
        }
        return Responses.success_response("monthly inventory data ", data=(data_dict))


class DayWiseInventory(ModelViewSet):
    queryset = TgtDepoInventoryStk.objects.all()
    serializer_class = TgtDepoInventoryStkSerializer
    filter_backends = (DjangoFilterBackend,)
    filterset_class = TgtDepoInventoryStkFilter
    pagination_class = CustomPagination

    def get(self, request, *args, **kwargs):
        product_wise_inv_data_today = (
            self.filter_queryset(self.get_queryset())
        ).values()
        return Responses.success_response(
            "day wise inventory data ", data=product_wise_inv_data_today
        )


class TgtRakeLoadingViewSet(ModelViewSet):
    queryset = TgtRakeLoading.objects.all().order_by("placement_date")
    serializer_class = TgtRakeLoadingSerializer
    pagination_class = CustomPagination
    permission_classes = (IsHandlingAgent | IsRRCCement,)
    filterset_class = TgtRakeLoadingFilter
    lookup_field = "rake_id"
    filter_backends = [DjangoFilterBackend]

    def get_serializer_class(self):
        if self.request.method == "GET":
            return TgtRakeLoadingReadOnlySerializer
        return super().get_serializer_class()

    def get_queryset(self):
        user = self.request.user
        user_roles = user.roles.values_list("role_name", flat=True).distinct()
        queryset = super().get_queryset()
        if (
            self.get_serializer_class() == TgtRakeLoadingReadOnlySerializer
            and self.request.method == "GET"
        ):
            if UserRoleChoice.HA in user_roles:
                return (
                    super()
                    .get_queryset()
                    .filter(
                        loading_details__ship_to_depot__in=settings.USER_DEPOT_MAPPING.get(
                            user.email
                        )
                    )
                    .distinct()
                )
        filterset = TgtRakeLoadingFilter(self.request.GET, queryset=queryset)
        queryset = filterset.qs

        return queryset


class TgtRakeLoadingDetailsViewSet(ModelViewSet):
    queryset = TgtRakeLoadingDetails.objects.select_related("rake")
    serializer_class = TgtRakeLoadingDetailsSerializer
    permission_classes = (IsHandlingAgent | IsRRCCement,)
    filter_backends = (DjangoFilterBackend,)
    filterset_class = TgtRakeLoadingDetailsFilter
    pagination_class = CustomPagination
    lookup_field = "rld_id"

    def get_queryset(self):
        user = self.request.user
        user_roles = user.roles.values_list("role_name", flat=True).distinct()
        if UserRoleChoice.HA in user_roles:
            return (
                super()
                .get_queryset()
                .filter(ship_to_depot__in=settings.USER_DEPOT_MAPPING.get(user.email))
            )
        return super().get_queryset()

    def __get_tgt_rake_loading_details_query(self, query_string, query=Q()):
        return (
            self.filter_queryset(self.get_queryset())
            .filter(query)
            .values_list(query_string, flat=True)
            .distinct()
        )

    def dropdown(self, request, *args, **kwargs):
        data = {
            "rr_no": self.__get_tgt_rake_loading_details_query(
                "rr_no", Q(rr_no__isnull=False)
            ),
            "excise_invoice_no": self.__get_tgt_rake_loading_details_query(
                "excise_invoice_no", Q(excise_invoice_no__isnull=False)
            ),
            "siding_name": self.__get_tgt_rake_loading_details_query(
                "siding_name", Q(siding_name__isnull=False)
            ),
            "siding_code": self.__get_tgt_rake_loading_details_query(
                "siding_code", Q(siding_code__isnull=False)
            ),
        }
        return Response(data)


class TgtRakeChargesViewSet(ModelViewSet):
    queryset = TgtRakeCharges.objects.all()
    serializer_class = TgtRakeChargesSerializer
    pagination_class = CustomPagination
    lookup_field = "rc_id"


class TgtRakeUnloadingDetailsViewSet(ModelViewSet):
    queryset = TgtRakeUnloadingDetails.objects.select_related("rld")
    serializer_class = TgtRakeUnloadingDetailsSerializer
    permission_classes = (IsHandlingAgent | IsRRCCement,)
    filter_backends = (DjangoFilterBackend,)
    filterset_class = TgtRakeUnloadingDetailsFilter
    pagination_class = CustomPagination
    lookup_field = "rk_unload_id"

    def get_queryset(self):
        return (
            super()
            .get_queryset()
            .filter(
                rld__ship_to_depot__in=settings.USER_DEPOT_MAPPING.get(
                    self.request.user.email
                )
            )
        )


class TgtRakeDisposalsViewSet(ModelViewSet):
    queryset = TgtRakeDisposals.objects.all()
    serializer_class = TgtRakeDisposalsSerializer
    filter_backends = (DjangoFilterBackend,)
    filterset_fields = ("rd_rk_unload_id", "rk_unload")
    pagination_class = CustomPagination
    lookup_field = "rd_rk_unload_id"


class TgtDayWiseLiftingViewSet(ModelViewSet):
    queryset = TgtDayWiseLifting.objects.all().order_by("lifting_date")
    serializer_class = TgtDayWiseLiftingSerializer
    # pagination_class = CustomPagination
    lookup_field = "daywise_lifting_id"
    filterset_class = TgtDayWiseLiftingFilter

    def create(self, request, *args, **kwargs):
        lifting_date = request.data.get("lifting_date")
        existing_record = TgtDayWiseLifting.objects.filter(
            lifting_date=lifting_date
        ).first()
        if existing_record:
            return Responses.error_response(
                {"detail": "A record with the same listing date already exists."}
            )
        return super().create(request, *args, **kwargs)


class StockAvailabilityViewSet(GenericAPIView):
    def get(self, request):
        # plant = request.query_params.get("plant")
        depo = request.query_params.get("depo")
        date = request.query_params.get("date")
        stocks_transit = (
            RrcIsoStockTransfer.objects.filter(
                consignee=depo,
                tax_invoice_date__isnull=False,
                receipt_date__isnull=True,
            )
            .values("product")
            .annotate(opening_stock=Sum("shipped_qty"))
        )
        return Responses.success_response(
            "stock in transist data fetched successfully", data=stocks_transit
        )


class StockAvailabilityindepot(GenericAPIView):
    def get(self, request):
        depo = request.query_params.get("depo")
        date = datetime.strptime(request.query_params.get("date"), "%Y-%m-%d").date()
        try:
            inventory_stk_queryset = (
                TgtDepoInventoryStk.objects.filter(whse_code=depo, trans_date=date)
                .values("item")
                .annotate(opening_stock=Sum("opening_stock"))
            )
        except:
            inventory_stk_queryset = []
        try:
            stock_received_todays = (
                TgtMrnData.objects.filter(
                    customer__startswith=depo, receipt_date__date=date
                )
                .values("ordered_item")
                .annotate(opening_stock=Sum("quantity_shipped"))
            )
            for i in stock_received_todays:
                i["item"] = i.pop("ordered_item")

        except:
            stock_received_todays = []
        result = {}
        add = []
        for items in list(inventory_stk_queryset) + list(stock_received_todays):
            key = items["item"]
            value = items["opening_stock"]
            if key in result:
                result[key] += value
            else:
                result[key] = value
        add.append(result)
        try:
            stock_dispatch_today = (
                TgtDepoDispatchData.objects.filter(
                    excise_invoice_no__excise_invoice_no__startswith=depo,
                    tax_invoice_date__date=date,
                )
                .values("product")
                .annotate(opening_stock=Sum("ordered_quantity"))
            )
            for i in stock_dispatch_today:
                i["item"] = i.pop("product")
        except:
            stock_dispatch_today = []
        # stocks_transit = TgtMrnData.objects.filter(
        #     excise_invoice_no__startswith=depo
        # ).aggregate(Sum("ordered_quantity"))
        result = {}
        change_keys = []
        dispatch_data = []
        for item in add:
            for k, v in item.items():
                new = {}
                new["item"] = k
                new["opening_stock"] = v
                change_keys.append(new)
        for item in change_keys:
            key = item["item"]
            value = item["opening_stock"]
            result[key] = value
        for item in list(stock_dispatch_today):
            key = item["item"]
            value = item["opening_stock"]
            if key in result:
                result[key] -= value
            else:
                result[key] = value
        dispatch_data.append(result)
        final_list = []

        for item in dispatch_data:
            for k, v in item.items():
                newss = {}
                newss["item"] = k
                newss["in_depot"] = v
                final_list.append(newss)
        return Responses.success_response(
            "stock in depot data fetched successfully", data=final_list
        )


class NewFreightInitiationStatusCountViewSet(GenericAPIView):
    def get(self, request):
        seven_day_before = datetime_date.today() - timedelta(days=7)
        status_count = (
            NewFreightInitiation.objects.filter(
                creation_date__date__gte=seven_day_before
            )
            .values("status")
            .annotate(Count("status"))
        )
        total_count = NewFreightInitiation.objects.filter(
            creation_date__date__gte=seven_day_before
        ).count()
        data_dict = {"status_count": status_count, "total_status_count": total_count}
        return Responses.success_response("data", data=data_dict)


class FreightChangeInitiationStatusCountViewSet(GenericAPIView):
    def get(self, request):
        seven_day_before = datetime_date.today() - timedelta(days=7)
        status_count = (
            FreightChangeInitiation.objects.filter(
                creation_date__date__gte=seven_day_before
            )
            .values("status")
            .annotate(Count("status"))
        )
        total_count = FreightChangeInitiation.objects.filter(
            creation_date__date__gte=seven_day_before
        ).count()
        data_dict = {"status_count": status_count, "total_status_count": total_count}
        return Responses.success_response("data", data=data_dict)


class FreightChangeInitiationDropdown(GenericAPIView):
    queryset = FreightChangeInitiation.objects.all()
    serializer_class = FreightChangeInitiationSerializer
    filter_backends = (DjangoFilterBackend,)
    filterset_class = NewFreightInitiationFilter

    def __get_freight_initiation_dropdown(self, query_string):
        return (
            self.filter_queryset(self.get_queryset())
            .values_list(query_string, flat=True)
            .annotate(Count(query_string))
            .order_by(query_string)
        )

    def get(self, request, *args, **kwargs):
        data = {
            "plant": self.__get_freight_initiation_dropdown("plant"),
            "ship_state": self.__get_freight_initiation_dropdown("ship_state"),
            "ship_district": self.__get_freight_initiation_dropdown("ship_district"),
            "segment": self.__get_freight_initiation_dropdown("segment"),
            "ship_city": self.__get_freight_initiation_dropdown("ship_city"),
            "mode": self.__get_freight_initiation_dropdown("mode"),
            "pack_type": self.__get_freight_initiation_dropdown("pack_type"),
        }
        return Responses.success_response(
            "freight intitation dropdown data.", data=data
        )


class NewFreightInitiationDropdown(GenericAPIView):
    queryset = NewFreightInitiation.objects.all()
    serializer_class = NewFreightInitiationSerializer
    filter_backends = (DjangoFilterBackend,)
    filterset_class = NewFreightInitiationFilter

    def __get_new_freight_initiation_dropdown(self, query_string):
        return (
            self.filter_queryset(self.get_queryset())
            .values_list(query_string, flat=True)
            .annotate(Count(query_string))
            .order_by(query_string)
        )

    def get(self, request, *args, **kwargs):
        data = {
            "plant": self.__get_new_freight_initiation_dropdown("plant"),
            "ship_state": self.__get_new_freight_initiation_dropdown("ship_state"),
            "ship_district": self.__get_new_freight_initiation_dropdown(
                "ship_district"
            ),
            "segment": self.__get_new_freight_initiation_dropdown("segment"),
            "ship_city": self.__get_new_freight_initiation_dropdown("ship_city"),
            "mode": self.__get_new_freight_initiation_dropdown("mode"),
            "pack_type": self.__get_new_freight_initiation_dropdown("pack_type"),
        }
        return Responses.success_response(
            "new freight intitation dropdown data.", data=data
        )


class ApprovalsBaseViewSet(DownloadUploadViewSet):
    filter_backends = (DjangoFilterBackend,)
    # permission_classes = (IsLogisticsHead,)
    pagination_class = CustomPagination

    def perform_create(self, serializer):
        serializer.save(status=ApprovalStatusChoices.PENDING.value)


class FreightChangeInitiationViewSet(ApprovalsBaseViewSet, RequestsCountMixin):
    # route_master table's columns mapping info:
    # dispatch_type = primary_secondary, plant=whse, remove pack_type column
    # segment = nt_tr, freight_type = dispatch_type, pack_type = attribute3
    # current_freight = freight_amount, description = route_description
    queryset = FreightChangeInitiation.objects.all().order_by("id")
    serializer_class = FreightChangeInitiationSerializer
    filterset_class = FreightChangeInitiationFilter
    file_name = "freight_change_initiation"


class FreightChangeInitiationDropdown(GenericAPIView):
    queryset = FreightChangeInitiation.objects.all()
    filter_backends = (DjangoFilterBackend,)
    filterset_class = FreightChangeInitiationFilter

    def __get_freight_change_initiation_dropdown(self, query_string, query=Q()):
        return (
            self.filter_queryset(self.get_queryset())
            .filter(query)
            .values_list(query_string, flat=True)
            .annotate(Count(query_string))
            .order_by(query_string)
        )

    def get(self, request, *args, **kwargs):
        data = {
            "plant": self.__get_freight_change_initiation_dropdown(
                "plant", Q(plant__isnull=False)
            ),
            "ship_state": self.__get_freight_change_initiation_dropdown(
                "ship_state", Q(ship_state__isnull=False)
            ),
            "ship_district": self.__get_freight_change_initiation_dropdown(
                "ship_district", Q(ship_district__isnull=False)
            ),
            "ship_city": self.__get_freight_change_initiation_dropdown(
                "ship_city", Q(ship_city__isnull=False)
            ),
            "mode": self.__get_freight_change_initiation_dropdown(
                "mode", Q(mode__isnull=False)
            ),
            "segment": self.__get_freight_change_initiation_dropdown(
                "segment", Q(segment__isnull=False)
            ),
            "pack_type": self.__get_freight_change_initiation_dropdown(
                "pack_type", Q(pack_type__isnull=False)
            ),
        }
        return Responses.success_response(
            "freight-change-initiation-dropdown", data=data
        )


class NewFreightInitiationViewSet(ApprovalsBaseViewSet, RequestsCountMixin):
    queryset = NewFreightInitiation.objects.all().order_by("id")
    serializer_class = NewFreightInitiationSerializer
    filter_backends = (DjangoFilterBackend,)
    filterset_class = NewFreightInitiationFilter

    def post(self, request, *args, **kwargs):
        request.data._mutable = True
        request.data["related_doc"] = request.FILES.get("related_doc")
        for data in request.data:
            contribution = data.get("contribution")
            approval = data["approval_type"]
            queryset = ApprovalThreshold.objects.filter(
                approval=approval, min__lte=contribution, max__gte=contribution
            )

            personas = [entry["persona"] for entry in queryset.values("persona")]

            for persona in personas:
                obj = NewFreightInitiation(persona=persona)
                obj.save()

        serializer = self.serializer_class(data=request.data)
        if not serializer.is_valid(raise_exception=True):
            return Responses.error_response("some issue rise", data=serializer.errors)
        serializer.save()
        return Responses.success_response(
            "data inserted successfully", data=serializer.data
        )


class RailRoadFlagViewSet(GenericAPIView):
    def get(self, request):
        try:
            date = datetime.strptime(
                request.query_params.get("date"), "%Y-%m-%d"
            ).date()
        except:
            date = None
        product = request.query_params.get("product")
        depot = request.query_params.get("depot")
        month = request.query_params.get("month")
        year = request.query_params.get("year")

        if date:
            if product:
                godown_sale_date_qty = (
                    TgtDepoDispatchData.objects.filter(
                        tax_invoice_date__date=date,
                        excise_invoice_no__excise_invoice_no__startswith=depot,
                        product=product,
                    )
                    .values("product", "sale_type")
                    .annotate(Count("product"), Count("sale_type"), Sum("shipped_qty"))
                )

            else:
                # product_list = TgtDepoDispatchData.objects.filter(
                #         tax_invoice_date__date=date,
                #         excise_invoice_no__excise_invoice_no__startswith=depot,
                #     ).values_list("product",flat=True).distinct()
                # data_list=[]
                # for product in product_list:
                #     godown_sale_date_qty = TgtDepoDispatchData.objects.filter(
                #             tax_invoice_date__date=date,
                #             excise_invoice_no__excise_invoice_no__startswith=depot,product=product
                #         ).values("sale_type").annotate(Count("sale_type"),Sum("shipped_qty"))

                #     data_dict={
                #         product:godown_sale_date_qty
                #     }
                #     data_list.append(data_dict)

                godown_sale_date_qty = (
                    TgtDepoDispatchData.objects.filter(
                        tax_invoice_date__date=date,
                        excise_invoice_no__excise_invoice_no__startswith=depot,
                    )
                    .values("sale_type")
                    .annotate(Count("sale_type"), Sum("shipped_qty"))
                )

            return Responses.success_response(
                "rail road flag data", data=godown_sale_date_qty
            )
        else:
            # product_list = TgtDepoDispatchData.objects.filter(
            #         tax_invoice_date__date__month=month,
            #         tax_invoice_date__date__year=year,
            #         excise_invoice_no__excise_invoice_no__startswith=depot,
            #     ).values_list("product",flat=True).distinct()
            # data_list=[]
            # for product in product_list:
            #     godown_sale_date_qty = TgtDepoDispatchData.objects.filter(
            #             tax_invoice_date__date__month=month,
            #             tax_invoice_date__date__year=year,
            #             excise_invoice_no__excise_invoice_no__startswith=depot,product=product
            #         ).values("sale_type").annotate(Count("sale_type"),Sum("shipped_qty"))
            #     print(godown_sale_date_qty)
            #     data_dict={
            #         product:godown_sale_date_qty
            #     }
            #     data_list.append(data_dict)
            if product:
                godown_sale_date_qty = (
                    TgtDepoDispatchData.objects.filter(
                        tax_invoice_date__date__month=month,
                        tax_invoice_date__date__year=year,
                        excise_invoice_no__excise_invoice_no__startswith=depot,
                        product=product,
                    )
                    .values("product", "sale_type")
                    .annotate(Count("product"), Count("sale_type"), Sum("shipped_qty"))
                )
            else:
                godown_sale_date_qty = (
                    TgtDepoDispatchData.objects.filter(
                        tax_invoice_date__date__month=month,
                        tax_invoice_date__date__year=year,
                        excise_invoice_no__excise_invoice_no__startswith=depot,
                    )
                    .values("sale_type")
                    .annotate(Count("sale_type"), Sum("shipped_qty"))
                )
            return Responses.success_response(
                "rail road flag data", data=godown_sale_date_qty
            )


class RailRoadFlagDropdownViewSet(GenericAPIView):
    def get(self, request):
        try:
            date = datetime.strptime(
                request.query_params.get("date"), "%Y-%m-%d"
            ).date()
        except:
            date = None
        depot = request.query_params.get("depot")
        month = request.query_params.get("month")
        year = request.query_params.get("year")
        if date:
            product_list = (
                TgtDepoDispatchData.objects.filter(
                    tax_invoice_date__date=date,
                    excise_invoice_no__excise_invoice_no__startswith=depot,
                )
                .values_list("product", flat=True)
                .distinct()
            )
        else:
            product_list = (
                TgtDepoDispatchData.objects.filter(
                    tax_invoice_date__date__month=month,
                    tax_invoice_date__date__year=year,
                    excise_invoice_no__excise_invoice_no__startswith=depot,
                )
                .values_list("product", flat=True)
                .distinct()
            )
        return Responses.success_response(
            "rail road flag product dropdown", data=product_list
        )

    # class GodownDispatchViewSet(GenericAPIView):
    #     def get(self, request):
    # month = datetime.today().month
    # year = datetime.today().year
    # try:
    #     date = datetime.strptime(request.query_params.get("date"), "%Y-%m-%d").date()
    # except:
    #     date = datetime_date.today()
    # month = 9
    # year = 2022
    # try:
    #     date = datetime.strptime(request.query_params.get("date"), "%Y-%m-%d").date()
    # except:
    #     date = None
    # depot = request.query_params.get("depot")
    # month =request.query_params.get("month")
    # year =request.query_params.get("year")
    # if date:
    #     product_list = TgtDepoDispatchData.objects.filter(
    #             tax_invoice_date__date=date,
    #             excise_invoice_no__excise_invoice_no__startswith=depot,
    #         ).values_list("product",flat=True).distinct()
    #     data_list=[]
    #     for product in product_list:
    #         godown_sale_date_qty = TgtDepoDispatchData.objects.filter(
    #                 tax_invoice_date__date=date,
    #                 excise_invoice_no__excise_invoice_no__startswith=depot,product=product
    #             ).values("sale_type").annotate(Count("sale_type"),Sum("shipped_qty"))
    #         data_dict={
    #             product:godown_sale_date_qty
    #         }
    #         data_list.append(data_dict)
    #     return Responses.success_response("ooooo",data=data_list)
    # else:
    #     product_list = TgtDepoDispatchData.objects.filter(
    #             tax_invoice_date__date__month=month,
    #             tax_invoice_date__date__year=year,
    #             excise_invoice_no__excise_invoice_no__startswith=depot,
    #         ).values_list("product",flat=True).distinct()
    #     data_list=[]
    #     for product in product_list:
    #         godown_sale_date_qty = TgtDepoDispatchData.objects.filter(
    #                 tax_invoice_date__date__month=month,
    #                 tax_invoice_date__date__year=year,
    #                 excise_invoice_no__excise_invoice_no__startswith=depot,product=product
    #             ).values("sale_type").annotate(Count("sale_type"),Sum("shipped_qty"))
    #         data_dict={
    #             product:godown_sale_date_qty
    #         }
    #         data_list.append(data_dict)
    #     return Responses.success_response("ooooo",data=data_list)
    """""" """"""
    # sale_type = ["GD", "GD-RK"]
    # if get_by_curr_data == "Daily":
    #     all_today_quantity_sum = TgtDepoDispatchData.objects.filter(
    #         tax_invoice_date__date=date,
    #         excise_invoice_no__excise_invoice_no__startswith=depot,
    #     ).aggregate(shipped_qty=Sum("shipped_qty"))

    #     try:
    #         total_qty = all_today_quantity_sum["shipped_qty"]
    #     except:
    #         total_qty = 0

    #     godown_sale_date_qty = TgtDepoDispatchData.objects.filter(
    #         sale_type=sale_type,
    #         tax_invoice_date__date=date,
    #         excise_invoice_no__excise_invoice_no__startswith=depot,
    #     ).aggregate(shipped_qty=Sum("shipped_qty"))

    #     try:
    #         godown_sale_date_qty = godown_sale_date_qty["shipped_qty"]
    #     except:
    #         godown_sale_date_qty = 0

    #     try:
    #         percentage = round(
    #             godown_sale_date_qty / total_qty * 100,
    #             2,
    #         )
    #     except:
    #         percentage = 0

    #     data = {
    #         "godown_sale_today_date_qty": godown_sale_date_qty,
    #         "total_qty": total_qty,
    #         "percentage": percentage,
    #     }

    #     return Responses.success_response("data", data=data)
    # else:
    #     all_monthly_quantity_sum = TgtDepoDispatchData.objects.filter(
    #         tax_invoice_date__date__month=month,
    #         tax_invoice_date__date__year=year,
    #         excise_invoice_no__excise_invoice_no__startswith=depot,
    #     ).aggregate(shipped_qty=Sum("shipped_qty"))
    #     try:
    #         total_qty = all_monthly_quantity_sum["shipped_qty"]
    #     except:
    #         total_qty = 0

    #     godown_sale_monthly_qty = TgtDepoDispatchData.objects.filter(
    #         sale_type__in=sale_type,
    #         tax_invoice_date__date__month=month,
    #         tax_invoice_date__date__year=year,
    #         excise_invoice_no__excise_invoice_no__startswith=depot,
    #     ).aggregate(shipped_qty=Sum("shipped_qty"))
    #     try:
    #         godown_sale_monthly_qty = godown_sale_monthly_qty["shipped_qty"]
    #     except:
    #         godown_sale_monthly_qty = 0
    #     try:
    #         percentage = round(
    #             godown_sale_monthly_qty / total_qty * 100,
    #             2,
    #         )
    #     except:
    #         percentage = 0
    #     data = {
    #         "godown_sale_monthly_qty": godown_sale_monthly_qty,
    #         "total_qty": total_qty,
    #         "percentage": percentage,
    #     }
    #     return Responses.success_response("data", data=data)


class CrossingDataViewSet(GenericAPIView):
    def get(self, request):
        month = datetime.today().month
        year = datetime.today().year
        month = 9
        year = 2022
        try:
            date = datetime.strptime(
                request.query_params.get("date"), "%Y-%m-%d"
            ).date()
        except:
            date = datetime_date.today()
        get_by_curr_data = request.query_params.get("get_by_curr_data")
        depot = request.query_params.get("depot")
        sale_type = ["RK", "GD-RK"]

        if get_by_curr_data == "Daily":
            all_quantity_sum = (
                TgtDepoDispatchData.objects.filter(
                    tax_invoice_date__date=date,
                    excise_invoice_no__excise_invoice_no__startswith=depot,
                )
                .values("shipped_qty")
                .aggregate(shipped_qty=Sum("shipped_qty"))
            )
            try:
                all_quantity_sum = all_quantity_sum["shipped_qty"]
            except:
                all_quantity_sum = 0
            rail_quantity_data = TgtDepoDispatchData.objects.filter(
                sale_type=sale_type,
                tax_invoice_date__date=date,
                excise_invoice_no__excise_invoice_no__startswith=depot,
            ).aggregate(shipped_qty=Sum("shipped_qty"))
            try:
                rail_quantity_data = rail_quantity_data["shipped_qty"]
            except:
                rail_quantity_data = 0

            td_quantity_data = TgtDepoDispatchData.objects.filter(
                sale_type="TP",
                tax_invoice_date__date=date,
                excise_invoice_no__excise_invoice_no__startswith=depot,
            ).aggregate(shipped_qty=Sum("shipped_qty"))
            try:
                td_quantity_data = td_quantity_data["shipped_qty"]
            except:
                td_quantity_data = 0
            try:
                road_qty = all_quantity_sum - rail_quantity_data
            except:
                road_qty = 0
            try:
                percentage = round(
                    (road_qty / td_quantity_data) * 100,
                    2,
                )
            except:
                percentage = 0

            data = {
                "crossing_sale_today_date_qty_data": td_quantity_data,
                "road_qty": road_qty,
                "percentage": percentage,
            }
            return Responses.success_response("data", data=data)
        else:
            all_monthly_quantity_sum = TgtDepoDispatchData.objects.filter(
                tax_invoice_date__date__month=month,
                tax_invoice_date__date__year=year,
                excise_invoice_no__excise_invoice_no__startswith=depot,
            ).aggregate(shipped_qty=Sum("shipped_qty"))
            try:
                all_monthly_quantity_sum = all_monthly_quantity_sum["shipped_qty"]

            except:
                all_monthly_quantity_sum = 0

            rail_quantity_data = TgtDepoDispatchData.objects.filter(
                sale_type__in=sale_type,
                tax_invoice_date__date__month=month,
                tax_invoice_date__date__year=year,
                excise_invoice_no__excise_invoice_no__startswith=depot,
            ).aggregate(shipped_qty=Sum("shipped_qty"))
            try:
                rail_quantity_data = rail_quantity_data["shipped_qty"]
            except:
                rail_quantity_data = 0

            td_quantity_data = TgtDepoDispatchData.objects.filter(
                tax_invoice_date__date__month=month,
                tax_invoice_date__date__year=year,
                sale_type="TP",
                excise_invoice_no__excise_invoice_no__startswith=depot,
            ).aggregate(shipped_qty=Sum("shipped_qty"))
            try:
                td_quantity_data = td_quantity_data["shipped_qty"]
            except:
                td_quantity_data = 0
            try:
                road_qty = all_monthly_quantity_sum - rail_quantity_data
            except:
                road_qty = 0
            try:
                percentage = round(
                    (td_quantity_data / road_qty) * 100,
                    2,
                )
            except:
                percentage = 0

            data = {
                "crossing_sale_monthly_qty_data": td_quantity_data,
                "road_qty": road_qty,
                "percentage": percentage,
            }
            return Responses.success_response("data", data=data)


class OrderStatusViewSet(GenericAPIView):
    def get(self, request):
        depo = request.query_params.get("depo")
        order_pendency_count = TgtSlhOrderPendency.objects.filter(
            auto_tagged_source=depo
        ).count()
        order_quantity_sum = TgtSlhOrderPendency.objects.filter(
            auto_tagged_source=depo
        ).aggregate(order_quantity=Sum("order_quantity"))
        final_data = {
            "order_pendency_count": order_pendency_count,
            "order_quantity_sum": order_quantity_sum["order_quantity"],
        }
        return Responses.success_response("data fetched", data=final_data)


class DiversionsViewSet(GenericAPIView):
    def get(self, request):
        month = datetime.today().month
        year = datetime.today().year
        try:
            date = datetime.strptime(
                request.query_params.get("date"), "%Y-%m-%d"
            ).date()
        except:
            date = datetime_date.today()
        month = 9
        year = 2022
        sale_type = ["RK", "GD-RK"]
        get_by_curr_data = request.query_params.get("get_by_month")
        depot = request.query_params.get("depot")
        if get_by_curr_data == "Daily":
            all_quantity_sum = (
                TgtDepoDispatchData.objects.filter(
                    tax_invoice_date__date=date,
                    excise_invoice_no__excise_invoice_no__startswith=depot,
                )
                .values("shipped_qty")
                .aggregate(shipped_qty=Sum("shipped_qty"))
            )
            try:
                all_quantity_sum = all_quantity_sum["shipped_qty"]
            except:
                all_quantity_sum = 0

            rail_quantity_data = TgtDepoDispatchData.objects.filter(
                sale_type__in=sale_type,
                tax_invoice_date__date=date,
                excise_invoice_no__excise_invoice_no__startswith=depot,
            ).aggregate(shipped_qty=Sum("shipped_qty"))

            try:
                rail_quantity_data = rail_quantity_data["shipped_qty"]
            except:
                rail_quantity_data = 0

            diversion_qty_data = TgtDepoDispatchData.objects.filter(
                sale_type="RD",
                tax_invoice_date__date=date,
                excise_invoice_no__excise_invoice_no__startswith=depot,
            ).aggregate(shipped_qty=Sum("shipped_qty"))

            try:
                diversion_qty_data = diversion_qty_data["shipped_qty"]
            except:
                diversion_qty_data = 0
            try:
                road_qty = all_quantity_sum - rail_quantity_data
            except:
                road_qty = 0
            try:
                percentage = round(
                    (diversion_qty_data / road_qty) * 100,
                    2,
                )
            except:
                percentage = 0

            data = {
                "diversion_sale_today_date_qty_data": diversion_qty_data,
                "road_qty": road_qty,
                "percentage": percentage,
            }
            return Responses.success_response("data", data=data)
        else:
            all_monthly_quantity_sum = TgtDepoDispatchData.objects.filter(
                tax_invoice_date__date__year=year,
                tax_invoice_date__date__month=month,
                excise_invoice_no__excise_invoice_no__startswith=depot,
            ).aggregate(shipped_qty=Sum("shipped_qty"))

            try:
                all_monthly_quantity_sum = all_monthly_quantity_sum["shipped_qty"]
            except:
                all_monthly_quantity_sum = 0
            rail_quantity_data = TgtDepoDispatchData.objects.filter(
                sale_type__in=sale_type,
                tax_invoice_date__date__year=year,
                tax_invoice_date__date__month=month,
                excise_invoice_no__excise_invoice_no__startswith=depot,
            ).aggregate(shipped_qty=Sum("shipped_qty"))
            try:
                rail_quantity_data = rail_quantity_data["shipped_qty"]
            except:
                rail_quantity_data = 0

            diversion_qty_data = TgtDepoDispatchData.objects.filter(
                tax_invoice_date__date__month=month,
                tax_invoice_date__date__year=year,
                sale_type="RD",
                excise_invoice_no__excise_invoice_no__startswith=depot,
            ).aggregate(shipped_qty=Sum("shipped_qty"))
            try:
                diversion_qty_data = diversion_qty_data["shipped_qty"]
            except:
                diversion_qty_data = 0
            try:
                road_qty = all_monthly_quantity_sum - rail_quantity_data
            except:
                road_qty = 0
            try:
                percentage = round(
                    (diversion_qty_data / road_qty) * 100,
                    2,
                )
            except:
                percentage = 0
            data = {
                "diversion_sale_monthly_qty_data": diversion_qty_data,
                "road_qty": road_qty,
                "percentage": percentage,
            }
            return Responses.success_response("data", data=data)


class RhFiringViewSet(GenericAPIView):
    def get(self, request):
        month = datetime.today().month
        year = datetime.today().year
        try:
            date = datetime.strptime(
                request.query_params.get("date"), "%Y-%m-%d"
            ).date()
        except:
            date = datetime_date.today()
        month = 9
        year = 2022
        get_by_curr_data = request.query_params.get("get_by_month")
        depot = request.query_params.get("depot")
        sale_type = ["RK", "GD-RK"]
        if get_by_curr_data == "Daily":
            rail_quantity_data = TgtDepoDispatchData.objects.filter(
                sale_type__in=sale_type,
                tax_invoice_date__date=date,
                excise_invoice_no__excise_invoice_no__startswith=depot,
            ).aggregate(shipped_qty=Sum("shipped_qty"))
            try:
                rail_quantity_data = rail_quantity_data["shipped_qty"]
            except:
                rail_quantity_data = 0

            rh_firing_qty_data = TgtDepoDispatchData.objects.filter(
                sale_type="RK",
                tax_invoice_date__date=date,
                excise_invoice_no__excise_invoice_no__startswith=depot,
            ).aggregate(shipped_qty=Sum("shipped_qty"))

            try:
                rh_firing_qty_data = rh_firing_qty_data["shipped_qty"]
            except:
                rh_firing_qty_data = 0
            try:
                percentage = round((rh_firing_qty_data / rail_quantity_data) * 100, 2)
            except:
                percentage = 0
            data = {
                "rail_sale_today_date_qty_data": rail_quantity_data,
                "rh_firing_qty_data": rh_firing_qty_data,
                "percentage": percentage,
            }
            return Responses.success_response("data", data=data)
        else:
            rail_quantity_data = TgtDepoDispatchData.objects.filter(
                sale_type__in=sale_type,
                tax_invoice_date__date__year=year,
                tax_invoice_date__date__month=month,
                excise_invoice_no__excise_invoice_no__startswith=depot,
            ).aggregate(shipped_qty=Sum("shipped_qty"))

            try:
                rail_quantity_data = rail_quantity_data["shipped_qty"]
            except:
                rail_quantity_data = 0
            rh_firing_qty_data = TgtDepoDispatchData.objects.filter(
                tax_invoice_date__date__year=year,
                tax_invoice_date__date__month=month,
                sale_type="RK",
                excise_invoice_no__excise_invoice_no__startswith=depot,
            ).aggregate(shipped_qty=Sum("shipped_qty"))

            try:
                rh_firing_qty_data = rh_firing_qty_data["shipped_qty"]
            except:
                rh_firing_qty_data = 0
            try:
                percentage = round(rh_firing_qty_data / rail_quantity_data * 100, 2)
            except:
                percentage = 0
            data = {
                "rail_sale_monthly_qty_data": rail_quantity_data,
                "rh_firing_qty_data": rh_firing_qty_data,
                "percentage": percentage,
            }
            return Responses.success_response("data", data=data)


class TgtSlhOrderPendencyQuantitySumAndCount(GenericAPIView):
    queryset = TgtSlhOrderPendency.objects.all()
    filter_backends = (DjangoFilterBackend,)
    filterset_class = TgtSlhOrderPendencyFilter

    def get(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())
        return Response(
            {
                "total_quantity": queryset.aggregate(Sum("order_quantity"))[
                    "order_quantity__sum"
                ],
                "total_entries": queryset.count(),
            }
        )


class TgtSlhServiceLevelDepoQuantitySumAndCount(GenericAPIView):
    queryset = TgtSlhServiceLevelDepo.objects.all()
    filter_backends = (DjangoFilterBackend,)
    filterset_class = TgtSlhServiceLevelDepoFilter

    def get(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())
        return Response(
            {
                "total_quantity": queryset.aggregate(Sum("dispatched_qty"))[
                    "dispatched_qty__sum"
                ],
                "total_entries": queryset.count(),
            }
        )


class OrderStatusRecievedViewSet(ModelViewSet):
    def get(self, request):
        depo = request.query_params.get("depo")
        month = request.query_params.get("month")
        year = request.query_params.get("year")
        date = request.query_params.get("date")
        # date = datetime.strptime(request.query_params.get("date"), "%Y-%m-%d").date()
        cement_list = TOebsFndLookupValues.objects.filter(
            lookup_type="SCL_CEMENT_ITEMS", enabled_flag="Y", zd_edition_name="SET2"
        ).values_list("meaning", flat=True)
        if date:
            delivery_id = (
                (
                    TgtMrnData.objects.filter(
                        customer__startswith=depo,
                        receipt_date__date=date,
                        ordered_item__in=cement_list,
                    )
                )
                .values("delivery_id")
                .distinct()
            )

            stock_received_todays = (
                TgtMrnData.objects.filter(
                    customer__startswith=depo,
                    ordered_item__in=cement_list,
                    delivery_id__in=delivery_id,
                )
                .values("ordered_item")
                .annotate(Count("ordered_item"), Sum("quantity_shipped"))
            )
        else:
            delivery_id = (
                (
                    TgtMrnData.objects.filter(
                        customer__startswith=depo,
                        receipt_date__date__month=month,
                        receipt_date__date__year=year,
                        ordered_item__in=cement_list,
                    )
                )
                .values("delivery_id")
                .distinct()
            )
            stock_received_todays = (
                TgtMrnData.objects.filter(
                    customer__startswith=depo,
                    ordered_item__in=cement_list,
                    delivery_id__in=delivery_id,
                )
                .values("ordered_item")
                .annotate(Count("ordered_item"), Sum("quantity_shipped"))
            )

        return Responses.success_response(
            "data fetched by successfully ", data=stock_received_todays
        )


class DepotOperationsViewSet(ModelViewSet):
    def get(self, request):
        depo = request.query_params.get("depo")
        date = request.query_params.get("date")
        month = request.query_params.get("month")
        year = request.query_params.get("year")
        cement_list = TOebsFndLookupValues.objects.filter(
            lookup_type="SCL_CEMENT_ITEMS", enabled_flag="Y", zd_edition_name="SET2"
        ).values_list("meaning", flat=True)
        if date:
            delevery_id = (
                TgtMrnData.objects.filter(
                    customer__startswith=depo,
                    actual_departure_date__isnull=False,
                    receipt_date__isnull=True,
                    ordered_item__in=cement_list,
                    excise_invoice_date__date=date,
                )
                .values("delivery_id")
                .distinct()
            )
            stocks_transit = (
                TgtMrnData.objects.filter(
                    customer__startswith=depo,
                    actual_departure_date__isnull=False,
                    receipt_date__isnull=True,
                    ordered_item__in=cement_list,
                    delivery_id__in=delevery_id,
                )
                .values("ordered_item")
                .annotate(Count("ordered_item"), Sum("ordered_quantity"))
            )
        else:
            delevery_id = (
                TgtMrnData.objects.filter(
                    customer__startswith=depo,
                    actual_departure_date__isnull=False,
                    receipt_date__isnull=True,
                    ordered_item__in=cement_list,
                    excise_invoice_date__date__month=month,
                    excise_invoice_date__date__year=year,
                )
                .values("delivery_id")
                .distinct()
            )
            stocks_transit = (
                TgtMrnData.objects.filter(
                    customer__startswith=depo,
                    actual_departure_date__isnull=False,
                    receipt_date__isnull=True,
                    ordered_item__in=cement_list,
                    delivery_id__in=delevery_id,
                )
                .values("ordered_item")
                .annotate(Count("ordered_item"), Sum("ordered_quantity"))
            )

        return Responses.success_response(
            "stock in transist data fetched successfully", data=stocks_transit
        )


class ProductWiseInventory(ModelViewSet):
    def get(self, request):
        try:
            date = datetime.strptime(
                request.query_params.get("date"), "%Y-%m-%d"
            ).date()
        except:
            date = datetime_date.today()
        # start_date = date - timedelta(days=59)
        depo = request.query_params.get("depo")
        product_wise_data = (
            TgtDepoInventoryStk.objects.filter(trans_date=date, whse_code=depo)
            .values("trans_date", "item", "location")
            .order_by("trans_date")
            .annotate(Sum("opening_stock"))
        )
        return Responses.success_response(
            "data fetched successfully", data=product_wise_data
        )


class ProductWiseDropdown(ModelViewSet):
    def get(self, request):
        try:
            date = datetime.strptime(
                request.query_params.get("date"), "%Y-%m-%d"
            ).date()
        except:
            date = datetime_date.today()
        start_date = date - timedelta(days=59)
        depo = request.query_params.get("depo")
        product_wise_data_dropdown = (
            TgtDepoInventoryStk.objects.distinct()
            .filter(trans_date__range=[start_date, date], whse_code=depo)
            .values_list("item", flat=True)
        )
        return Responses.success_response(
            "products name dropdown", data=product_wise_data_dropdown
        )


class MonthlyWiseInventory(ModelViewSet):
    def get(self, request, *args, **kwargs):
        depo = request.query_params.get("depo")
        try:
            selected_date = datetime.strptime(
                request.query_params.get("date"), "%Y-%m-%d"
            ).date()
        except:
            selected_date = datetime_date.today()
        start_date = selected_date - timedelta(weeks=8)

        weeks = (
            TgtDepoInventoryStk.objects.filter(
                trans_date__gte=start_date,
                trans_date__lte=selected_date,
                whse_code=depo,
            )
            .annotate(
                week_number=ExtractWeek("trans_date"),
                month_name=ExtractMonth("trans_date", "F"),
            )
            .order_by("week_number")
            .values_list("week_number", "month_name")
            .distinct()
        )

        final_list = []
        for week in weeks:
            items = (
                TgtDepoInventoryStk.objects.filter(
                    trans_date__gte=start_date,
                    trans_date__lte=selected_date,
                    whse_code=depo,
                )
                .annotate(week_number=ExtractWeek("trans_date"))
                .filter(week_number=week[0])
                .values_list("item", flat=True)
            ).distinct()
            items_data_list = []
            for item in items:
                result = (
                    TgtDepoInventoryStk.objects.filter(
                        trans_date__gte=start_date,
                        trans_date__lte=selected_date,
                        whse_code=depo,
                        item=item,
                    )
                    .annotate(week_number=ExtractWeek("trans_date"))
                    .filter(week_number=week[0])
                    .values("week_number", "item", "location")
                    .order_by("week_number", "item")
                    .annotate(Count("item"), Count("location"), (Sum("opening_stock")))
                )
                cut = 0.00
                accept = 0.00
                damage = 0.00
                shortage = 0.00
                for item in result:
                    if item["location"] == "ACCEPT":
                        accept = item["opening_stock__sum"]

                    if item["location"] == "CUT":
                        cut = item["opening_stock__sum"]

                    if item["location"] == "DAMAGE":
                        damage = item["opening_stock__sum"]

                    if item["location"] == "SHORTAGE":
                        shortage = item["opening_stock__sum"]
                average = round(
                    (
                        (float(accept) + float(cut) + float(damage) - float(shortage))
                        / 7
                    ),
                    2,
                )

                result = {"item": item["item"], "average": average}
                items_data_list.append(result)

            month = calendar.month_name[week[1]]

            data_dict = {
                "week": "week" + ":" + str(week[0]) + "," + str(month)[0:3],
                "items": items_data_list,
            }
            final_list.append(data_dict)

        return Responses.success_response(
            "get data week and month wise successfully", data=final_list
        )


class TgtPlantDispatchDataDropdownAPIView(GenericAPIView):
    queryset = TgtPlantDispatchData.objects.exclude(product="CLINKER")
    filter_backends = (DjangoFilterBackend,)
    # filterset_fields = ("di_so", "mode_of_transport")
    filterset_class = TgtPlantDispatchDataFilter

    def get(self, request, *args, **kwargs):
        return Response(
            {
                "excise_invoice_no": self.filter_queryset(self.get_queryset())
                .values_list("excise_invoice_no", flat=True)
                .annotate(Count("excise_invoice_no"))
            }
        )


# class EpodDataViewSet(ModelViewSet):
#     queryset = TgtPlantDispatchData.objects.all()
#     filter_backends = (DjangoFilterBackend,)
#     filterset_class = TgtPlantDispatchDataFilter

#     def get(self, request, *args, **kwarg):
#         epod_df = pd.DataFrame(
#             EpodData.objects.filter(truck_reach_time__isnull=False).values(
#                 "id", "excise_invoice_no", "delivery_id"
#             )
#         )
#         tgt_plant_dispatch_data = pd.DataFrame(
#             self.filter_queryset(self.get_queryset()).values(
#                 "product",
#                 "shipped_qty",
#                 "truck_type",
#                 "vehicle_no",
#                 "freight_term",
#                 "token_no",
#                 "cust_category",
#                 "cust_sub_category",
#                 "customer_name",
#                 "consignee",
#                 "ship_city",
#                 "ship_taluka",
#                 "ship_district",
#                 "ship_state",
#                 "brand",
#                 "delivery_id",
#             )
#         )
#         merge_df = pd.merge(
#             epod_df, tgt_plant_dispatch_data, on="delivery_id", how="left"
#         ).fillna("")

#         data = merge_df.to_dict(orient="records")

#         return Response(data)

#     def patch(self, request):
#         id = request.query_params.get("id")
#         epod_obj = EpodData.objects.get(id=id)
#         epod_obj.truck_reach_time = datetime.now()
#         serializer = EpodDataSerializer(epod_obj, data=request.data, partial=True)
#         if serializer.is_valid():
#             serializer.save(
#                 last_updated_by=request.user.id, last_update_login=request.user.id
#             )
#             return Responses.success_response(
#                 "data updated by id successfully ", data=serializer.data
#             )
#         else:
#             return Responses.error_response("update failed", data=serializer.errors)


class EpodDataViewSet(ModelViewSet):
    # end_date = datetime_date.today()
    # start_date = end_date - timedelta(days=30)
    # delivery_ids = TgtPlantDispatchData.objects.filter(
    #     tax_invoice_date__date__range=[start_date, end_date]
    # ).values_list("delivery_id", flat=True)
    # queryset = EpodData.objects.filter(
    #     delivery_id__in=delivery_ids,
    # )
    queryset = EpodData.objects.all()

    serializer_class = EpodDataSerializer
    filter_backends = (DjangoFilterBackend, SearchFilter)
    filterset_class = EpodDataFilter
    search_fields = ("delivery_id",)
    pagination_class = CustomPagination
    lookup_field = "id"

    def get(self, requset):
        end_date = datetime_date.today()
        start_date = end_date - timedelta(days=30)
        email = requset.user.email
        transporter_name = TgtRlsRoleData.objects.filter(email=email).values_list(
            "transporter_company_name", flat=True
        )
        delivery_ids = TgtPlantDispatchData.objects.filter(
            tax_invoice_date__date__range=[start_date, end_date],
            transporter__in=transporter_name,
        ).values_list("delivery_id", flat=True)

        queryset = self.filter_queryset(self.get_queryset()).filter(
            delivery_id__in=delivery_ids,
        )
        serializers = EpodDataSerializer(queryset, many=True)
        return self.get_paginated_response(self.paginate_queryset(serializers.data))


class TgtPlantDispatchDataAPIView(ListAPIView):
    queryset = (
        TgtPlantDispatchData.objects.filter(~Q(product="CLINKER"))
        .select_related("delivery_id")
        .order_by("-tax_invoice_date")
    )
    serializer_class = TgtPlantDispatchDataSerializer
    filter_backends = (DjangoFilterBackend,)
    filterset_class = TgtPlantDispatchDataFilter
    pagination_class = CustomPagination

    def get_serializer_class(self):
        class CustomTgtPlantDispatchDataSerializer(TgtPlantDispatchDataSerializer):
            def to_representation(self, instance):
                data = super().to_representation(instance)
                if data["brand"] == "Cemento":
                    data["brand"] = "Rockstrong"
                return data

        return CustomTgtPlantDispatchDataSerializer


class ReasonForFreightChangeDropdown(GenericAPIView):
    queryset = ReasonsForFreightChange.objects.all()
    filter_backends = (DjangoFilterBackend,)

    def __get_reason_for_freight_change_query(self, query_string):
        return (
            self.filter_queryset(self.get_queryset())
            .values_list(query_string, flat=True)
            .annotate(Count(query_string))
            .order_by(query_string)
        )

    def get(self, request, *args, **kwargs):
        data = {"reasons": self.__get_reason_for_freight_change_query("reasons")}

        return Responses.success_response(
            "reasons for freight change dropdown", data=data
        )


class ReasonsForDemurrageWharfageView(ModelViewSet):
    pagination_class = CustomPagination
    permission_classes = (IsAuthenticated,)

    def get(self, request):
        queryset = ReasonsForDemurrageWharfage.objects.filter(
            category=request.query_params.get("category"),
            rake_point_type=request.query_params.get("rake_point_type"),
        )
        serailizer = ReasonsForDemurrageWharfageSerializer(queryset, many=True)
        return Responses.success_response(
            "fetched data sucessfully", data=serailizer.data
        )


class EpodDataVehicleNoAndDeliveryIdDropdown(GenericAPIView):
    end_date = datetime_date.today()
    start_date = end_date - timedelta(days=3)
    queryset = TgtPlantDispatchData.objects.filter(
        tax_invoice_date__date__range=[start_date, end_date]
    )
    filter_backends = (DjangoFilterBackend, SearchFilter)
    filterset_class = TgtPlantDispatchDataFilter
    search_fields = (
        "delivery_id",
        "vehicle_no",
        "ship_state",
        "ship_district",
        "ship_taluka",
        "ship_city",
        "customer_name",
        "excise_invoice_no",
    )

    def __get_dropdown(self, query):
        queryset = (
            self.filter_queryset(self.get_queryset())
            .values_list(query, flat=True)
            .distinct()
        )
        return queryset

    def get(self, request, *args, **kwargs):
        data = {
            "delivery_id": self.__get_dropdown("delivery_id"),
            "vehicle_no": self.__get_dropdown("vehicle_no"),
            "ship_state": self.__get_dropdown("ship_state"),
            "ship_district": self.__get_dropdown("ship_district"),
            "ship_taluka": self.__get_dropdown("ship_taluka"),
            "ship_city": self.__get_dropdown("ship_city"),
            "consignee": self.__get_dropdown("consignee"),
            "customer_name": self.__get_dropdown("customer_name"),
            "excise_invoice_no": self.__get_dropdown("excise_invoice_no"),
            "erp_order_number": self.__get_dropdown("excise_invoice_no"),
            "crm_order_number": self.__get_dropdown("excise_invoice_no"),
        }
        return Responses.success_response(
            "tgt dispatch delivery id and vehicle no dropdown", data=data
        )


class TgtRakeLoadingGetViewSet(ModelViewSet):
    queryset = TgtRakeLoading.objects.all().order_by("placement_date")
    serializer_class = TgtRakeLoadingGetSerializer
    pagination_class = CustomPagination
    permission_classes = (IsHandlingAgent | IsRRCCement,)
    filterset_class = TgtRakeLoadingFilter
    lookup_field = "rake_id"
    filter_backends = [DjangoFilterBackend]


class TgtRakeLoadingDetailDropdown(GenericAPIView):
    queryset = TgtRakeLoadingDetails.objects.all()
    filter_backends = (DjangoFilterBackend,)
    filterset_class = TgtRakeLoadingDetailsFilter

    def __get_tgt_rake_loading_details__dropdown(
        self, query_string, total_depots, request
    ):
        depot = request.query_params["depot"].lower()
        queryset = (
            self.filter_queryset(self.get_queryset())
            .values_list(query_string, flat=True)
            .annotate(Count(query_string))
            .order_by(query_string)
        )

        if depot == "true":
            # user_email = request.user.email

            return (
                self.filter_queryset(self.get_queryset())
                .filter(ship_to_depot__in=total_depots)
                .values_list(query_string, flat=True)
                .annotate(Count(query_string))
                .order_by(query_string)
            )
        else:
            return (
                self.filter_queryset(self.get_queryset())
                # .filter(ship_to_depot__in=total_depots)
                .values_list(query_string, flat=True)
                .annotate(Count(query_string))
                .order_by(query_string)
            )

    def get(self, request, *args, **kwargs):
        total_depots = settings.USER_DEPOT_MAPPING.get(f"{request.user.email}", [])

        data = {
            "excise_invoice_no": self.__get_tgt_rake_loading_details__dropdown(
                "excise_invoice_no", total_depots, request
            ),  # type: ignore
            "rake_point": self.__get_tgt_rake_loading_details__dropdown(
                "rake_point", total_depots, request
            ),  # type: ignore
            "rr_no": self.__get_tgt_rake_loading_details__dropdown(
                "rr_no", total_depots, request
            ),
        }
        rake_id = request.query_params.get("rake_id")
        if rake_id is not None:
            data["rake_id"] = self.__get_tgt_rake_loading_details__dropdown(
                "rake_id", total_depots, request
            )

        return Responses.success_response(
            "tgt rake loading details dropdown", data=data
        )


class TgtRakeLoadingDropdown(GenericAPIView):
    queryset = TgtRakeLoading.objects.all()
    filter_backends = (DjangoFilterBackend,)
    filterset_class = TgtRakeLoadingFilter

    def __get_tgt_rake_loading_dropdown(self, query_string):
        return (
            self.filter_queryset(self.get_queryset())
            .values_list(query_string, flat=True)
            .annotate(Count(query_string))
            .order_by(query_string)
        )

    def get(self, request, *args, **kwargs):
        data = {
            "dispatch_from_plant": self.__get_tgt_rake_loading_dropdown(
                "dispatch_from_plant"
            ),
            "siding_code": self.__get_tgt_rake_loading_dropdown("siding_code"),
        }

        return Responses.success_response("tgt rake loading  dropdown", data=data)


class SumOfClosingPointView(ModelViewSet):
    def get(self, request):
        value = TgtDayWiseLifting.objects.filter(
            rk_unload=request.query_params.get("rk_unload_id")
        ).aggregate(closing_quantity_sum=Sum("closing_quantity"))
        return Responses.success_response(
            "day wise lifting closing quantity fetched succesfully", data=value
        )


class WharfageSlabsViewSet(DownloadUploadViewSet):
    queryset = WharfageSlabs.objects.all()
    serializer_class = WharfageSlabsSerializer
    pagination_class = CustomPagination
    file_name = "wharfage_slabs"


class DemurrageSlabsViewSet(DownloadUploadViewSet):
    queryset = DemurrageSlabs.objects.all()
    serializer_class = DemurrageSlabsSerializer
    pagination_class = CustomPagination
    file_name = "demurrage_slabs"


class CrwcChargesMasterViewSet(DownloadUploadViewSet):
    queryset = CrwcChargesMaster.objects.all()
    serializer_class = CrwcChargesMasterSerializer
    filter_backends = (DjangoFilterBackend,)
    filterset_fields = ("rake_point",)
    pagination_class = CustomPagination
    file_name = "crwc_charges_master"


class WaiverCommissionMasterViewSet(DownloadUploadViewSet):
    queryset = WaiverCommissionMaster.objects.all()
    serializer_class = WaiverCommissionMasterSerializer
    filter_backends = (DjangoFilterBackend,)
    filterset_fields = ("agent",)
    pagination_class = CustomPagination
    file_name = "waiver_commission_master"


class RailExpensesDetailsViewSet(DownloadUploadViewSet):
    queryset = RailExpensesDetails.objects.all()
    serializer_class = RailExpensesDetailsSerializer
    # pagination_class = CustomPagination
    filterset_class = RailExpensesDetailsFilterset
    file_name = "rail_expenses_details"

    # def post(self, request):
    def get_queryset(self):
        total_depots = settings.USER_DEPOT_MAPPING.get(f"{self.request.user.email}", [])
        if (
            self.get_serializer_class() == RailExpensesDetailsSerializer
            and self.request.method == "GET"
        ):
            queryset = RailExpensesDetails.objects.filter(
                ship_to_depot__in=total_depots
            ).distinct()
            return queryset

    def create(self, request):
        serializer = RailExpensesDetailsPostSerializer(
            data=request.data,
            context={
                "request_user": request.user.id,
            },
            many=True,
        )
        if serializer.is_valid(raise_exception=True):
            serializer.save()
            return Responses.success_response(
                "data saved succesfully", data=serializer.data
            )

        # request.data[0]["created_by"] = request.user.id
        # request.data[0]["last_updated_by"] = request.user.id
        # request.data[0]["last_update_login"] = request.user.id

        # rail_exp_serializer = RailExpensesDetailsPostSerializer(data=request.data[0])
        # if not rail_exp_serializer.is_valid(raise_exception=True):
        #     return Responses.error_response(
        #         "some issue rise", data=rail_exp_serializer.errors
        #     )
        # rail_exp_serializer.save()
        case_study_data = serializer.data
        rake_id = case_study_data["rake_id"]

        if rake_id:
            expense_queryset = RailExpensesDetails.objects.filter(
                rake_id=rake_id
            ).aggregate(
                rate_per_wagon_cal=ExpressionWrapper(
                    Sum("dm_amount_wth_gst")
                    / (Sum("rate_per_wagon") * Sum("dm_hours")),
                    output_field=DecimalField(),
                ),
                dm_hours_sum=Sum("dm_hours"),
                wagon_under_dm_average=Avg("wagon_under_dm"),
                total_dm_amount_sum=Sum("total_dm_amount"),
                total_gst_amount=Sum("total_gst_amount"),
            )
            rake_charges_dict = {
                "dm_hours": expense_queryset["dm_hours_sum"],
                "dm_rate_per_wagon": expense_queryset["rate_per_wagon_cal"],
                "dm_total_wgn_placed": expense_queryset["wagon_under_dm_average"],
                "dm_total_gst": expense_queryset["total_gst_amount"],
                "total_demurrage_amount": expense_queryset["total_dm_amount_sum"],
            }

            TgtRakeCharges.objects.filter(rld__rld_id=rake_id).update(
                **rake_charges_dict
            )

            return Responses.success_response("data updated", data=serializer.data)


class WaiverCommissionMasterDropdown(GenericAPIView):
    queryset = WaiverCommissionMaster.objects.all()
    filter_backends = (DjangoFilterBackend,)
    # filterset_class = WaiverCommissionMasterFilter

    def _get_waiver_commission_query(self, query_string):
        return (
            self.filter_queryset(self.get_queryset())
            .values_list(query_string, flat=True)
            .annotate(Count(query_string))
            .order_by(query_string)
        )

    def get(self, request, *args, **kwargs):
        data = {
            "agent": self._get_waiver_commission_query("agent"),
        }
        return Responses.success_response(
            "waiver commission master dropdown", data=data
        )


class CrwcChargesMasterDropdown(GenericAPIView):
    queryset = CrwcChargesMaster.objects.all()
    filter_backends = (DjangoFilterBackend,)
    # filterset_class = CrwcChargesMasterFilter

    def _get_crw_charges_master_query(self, query_string):
        return (
            self.filter_queryset(self.get_queryset())
            .values_list(query_string, flat=True)
            .annotate(Count(query_string))
            .order_by(query_string)
        )

    def get(self, request, *args, **kwargs):
        data = {
            "rake_point": self._get_crw_charges_master_query("rake_point"),
        }
        return Responses.success_response("crwc charges master dropdown", data=data)


class HourlyLiftingEfficiencyMasterViewSet(ModelViewSet):
    queryset = HourlyLiftingEfficiencyMaster.objects.all()
    serializer_class = HourlyLiftingEfficiencyMasterSerializer
    filter_backends = (DjangoFilterBackend,)
    pagination_class = CustomPagination
    lookup_field = "id"


class HourlyLiftingEfficiencyViewSet(ModelViewSet):
    queryset = HourlyLiftingEfficiency.objects.all()
    serializer_class = HourlyLiftingEfficiencySerializer
    filter_backends = (DjangoFilterBackend,)
    pagination_class = CustomPagination
    lookup_field = "id"


class SidingConstraintsViewSet(DownloadUploadViewSet):
    queryset = SidingConstraints.objects.all()
    serializer_class = SidingConstraintsSerializer
    filter_backends = (DjangoFilterBackend,)
    filterset_class = SidingConstraintsFilterset
    pagination_class = CustomPagination
    file_name = "siding_constraints_master"

    def create(self, request):
        serializer = SidingConstraintsSerializer(
            data=request.data,
            context={
                "request_user": request.user.id,
            },
            many=True,
        )
        if serializer.is_valid(raise_exception=True):
            serializer.save()
            return Responses.success_response(
                "data saved succesfully", data=serializer.data
            )
        return Responses.error_response("some issue occured", data=serializer.errors)


class CostsMasterDetailViewSet(ModelViewSet):
    """cost master detail class"""

    queryset = TgtRakeLoadingDetails.objects.all().order_by("rake_point")
    serializer_class = CostsMasterDetailSerializer
    permission_classes = (IsHandlingAgent,)
    filter_backends = (DjangoFilterBackend,)
    filterset_class = TgtRakeLoadingDetailsFilter
    pagination_class = CustomPagination
    lookup_field = "id"

    def get_queryset(self):
        user = self.request.user
        user_roles = user.roles.values_list("role_name", flat=True).distinct()
        if UserRoleChoice.HA in user_roles:
            return (
                super()
                .get_queryset()
                .filter(ship_to_depot__in=settings.USER_DEPOT_MAPPING.get(user.email))
            )
        return super().get_queryset()

    def get(self, request):
        queryset = self.filter_queryset(self.get_queryset())
        serializer = self.serializer_class(queryset, many=True)
        unique_results = self.remove_duplicates(serializer.data)
        return Responses.success_response(
            "data fetched successfully", data=unique_results
        )

    def remove_duplicates(self, data):
        seen = set()
        unique_results = []
        for item in data:
            unique_key = (
                item["rake_point"],
                item["packing_type"],
                item["ship_to_depot"],
            )
            if unique_key not in seen:
                seen.add(unique_key)
                unique_results.append(item)
        return unique_results


class CostMasterCostHeadViewSet(GenericAPIView):
    queryset = RakeHandlingMaster.objects.all()
    filter_backends = (DjangoFilterBackend,)
    filterset_fields = ("rake_point",)

    def get(self, request, *args, **kwargs):
        rake_handling_cost_paper = (
            self.filter_queryset(self.get_queryset())
            .filter(packing="PAPER")
            .aggregate(Avg("total_cost_per_mt"))
        )
        rake_handling_cost_hdpe = (
            self.filter_queryset(self.get_queryset())
            .filter(~Q(packing="PAPER"))
            .aggregate(Avg("total_cost_per_mt"))
        )
        avg_rh_firing_hdpe = DepoWiseFreightMaster.objects.filter(
            ~Q(pack_mat="PAPER"), sale_type="RK"
        ).aggregate(Avg("average_freight"))
        avg_rh_firing_paper = DepoWiseFreightMaster.objects.filter(
            pack_mat="PAPER", sale_type="RK"
        ).aggregate(Avg("average_freight"))
        return Responses.success_response(
            "data fetched successfully",
            data={
                "rake_handling_cost_paper": rake_handling_cost_paper[
                    "total_cost_per_mt__avg"
                ],
                "rake_handling_cost_hdpe": rake_handling_cost_hdpe[
                    "total_cost_per_mt__avg"
                ],
                "avg_rh_firing_hdpe": avg_rh_firing_hdpe["average_freight__avg"],
                "avg_rh_firing_paper": avg_rh_firing_paper["average_freight__avg"],
                "road_handling_crwc_hdpe": None,
                "road_handling_crwc_paper": None,
            },
        )


# class LiftingPatternViewSet(ModelViewSet):
#     queryset = LiftingPattern.objects.all()
#     serializer_class = LiftingPatternSerializer
#     filter_backends = (DjangoFilterBackend,)
#     pagination_class = CustomPagination
#     lookup_field = "id"

#     def create(self, request):
#         serializer = LiftingPatternSerializer(
#             data=request.data,
#             context={
#                 "request_user": request.user.id,
#             },
#             many=True,
#         )
#         if serializer.is_valid(raise_exception=True):
#             serializer.save()
#             return Responses.success_response(
#                 "data saved succesfully", data=serializer.data
#             )
#         return Responses.error_response("some issue occured", data=serializer.errors)


class LiftingPatternViewSet(DownloadUploadViewSet):
    permission_classes = (IsAuthenticated,)
    queryset = LiftingPattern.objects.all()
    serializer_class = LiftingPatternSerializer
    filter_backends = (DjangoFilterBackend,)
    # filterset_class = LiftingPatternFilter
    filterset_fields = (
        "run_id",
        "rake_id",
        "rake_point",
        "rake_point_code",
        "brand",
        "packaging",
    )
    # pagination_class = CustomPagination
    # sorting_field = "id"
    lookup_field = "id"


class GetTimeCombination(ModelViewSet):
    queryset = (
        TgtRakeLoadingDetails.objects.filter()
        .values("org_id", "packing_type")
        .annotate(total_qty_dispatch=Sum("qty_dispatch_frm_plant"))
        .order_by()
    )
    filter_backends = (DjangoFilterBackend,)
    filterset_class = TgtRakeLoadingDetailsFilter

    def get(self, request):
        queryset = self.filter_queryset(self.get_queryset())
        for data in queryset:
            if data["packing_type"] != "PAPER":
                data["packing_type"] = "HDPE"
            else:
                data["packing_type"]
        df = pd.DataFrame(queryset)
        result = (
            df.groupby(["org_id", "packing_type"])["total_qty_dispatch"]
            .apply(lambda x: x.astype(float).sum())
            .reset_index()
        )
        result = round(result, 2)
        result = result.to_dict(orient="records")

        return Responses.success_response("Data Fetched Successfully ", data=result)


class RailExpensesDetailsWfViewSet(DownloadUploadViewSet):
    queryset = RailExpensesDetailsWarfage.objects.all()
    serializer_class = RailExpensesDetailsWfSerializer
    # pagination_class = CustomPagination
    filterset_class = RailExpensesDetailsWfFilterset
    file_name = "rail_expenses_details"

    def get_queryset(self):
        total_depots = settings.USER_DEPOT_MAPPING.get(f"{self.request.user.email}", [])
        if (
            self.get_serializer_class() == RailExpensesDetailsWfSerializer
            and self.request.method == "GET"
        ):
            queryset = RailExpensesDetailsWarfage.objects.filter(
                ship_to_depot__in=total_depots
            ).distinct()
            return queryset

    # def post(self, request):
    #     request.data[0]["created_by"] = request.user.id
    #     request.data[0]["last_updated_by"] = request.user.id
    #     request.data[0]["last_update_login"] = request.user.id

    #     rail_exp_serializer = RailExpensesDetailsWfPostSerializer(data=request.data[0])
    #     if not rail_exp_serializer.is_valid(raise_exception=True):
    #         return Responses.error_response(
    #             "some issue rise", data=rail_exp_serializer.errors
    #         )
    #     rail_exp_serializer.save()
    #     case_study_data = rail_exp_serializer.data
    #     rake_id = case_study_data["rake_id"]
    def create(self, request):
        serializer = RailExpensesDetailsWfPostSerializer(
            data=request.data,
            context={
                "request_user": request.user.id,
            },
            many=True,
        )
        if serializer.is_valid(raise_exception=True):
            serializer.save()
            return Responses.success_response(
                "data saved succesfully", data=serializer.data
            )

        case_study_data = serializer.data
        rake_id = case_study_data["rake_id"]
        if rake_id:
            expense_queryset = RailExpensesDetailsWarfage.objects.filter(
                rake_id=rake_id
            ).aggregate(
                wf_rate_per_wagon=ExpressionWrapper(
                    Sum("wf_amount_wth_gst")
                    / (Sum("wf_rate_per_wagon") * Sum("wf_hours")),
                    output_field=DecimalField(),
                ),
                wf_hours_sum=Sum("wf_hours"),
                wagon_under_wf_average=Avg("wagon_under_wf"),
                total_wf_amount_sum=Sum("total_wf_amount"),
                total_wf_gst_amount=Sum("total_wf_gst_amount"),
            )
            rake_charges_dict = {
                "wf_hours": expense_queryset["wf_hours_sum"],
                "wf_rate_per_wagon": expense_queryset["wf_rate_per_wagon"],
                "wf_total_wgn_placed": expense_queryset["wagon_under_wf_average"],
                "wf_total_gst": expense_queryset["total_wf_gst_amount"],
                "total_wharfage_amount": expense_queryset["total_wf_amount_sum"],
            }

            TgtRakeCharges.objects.filter(rld__rld_id=rake_id).update(
                **rake_charges_dict
            )

            return Responses.success_response("data updated", data=serializer.data)


class EpodDataReachedViewSet(ModelViewSet):
    end_date = datetime_date.today()
    start_date = end_date - timedelta(days=30)
    delivery_ids = TgtPlantDispatchData.objects.filter(
        tax_invoice_date__date__range=[start_date, end_date]
    ).values_list("delivery_id", flat=True)

    queryset = EpodData.objects.filter(
        delivery_id__in=delivery_ids,
    )
    # serializer_class = EpodDataSerializer
    filter_backends = (DjangoFilterBackend, SearchFilter)
    filterset_class = EpodDataFilter
    search_fields = ("delivery_id",)
    pagination_class = CustomPagination
    lookup_field = "id"

    def get(self, request, *args, **kwargs):
        end_date = datetime_date.today()
        start_date = end_date - timedelta(days=30)

        seven_days_ago = end_date - timedelta(days=7)

        first_df = pd.DataFrame(
            self.filter_queryset(self.get_queryset())
            .filter(epod_completion_time__date__range=[seven_days_ago, end_date])
            .values()
        )
        second_df = pd.DataFrame(
            self.filter_queryset(self.get_queryset())
            .filter(epod_completion_time__isnull=True)
            .values()
        )

        result = pd.concat([first_df, second_df], ignore_index=True)

        third_df = pd.DataFrame(
            TgtPlantDispatchData.objects.filter(
                delivery_id__in=result["delivery_id"]
            ).values(
                "organization_id",
                "product",
                "tax_invoice_no",
                # "excise_invoice_no",
                "ordered_qty",
                "shipped_qty",
                "truck_type",
                "vehicle_no",
                "lr_gr_no",
                "lr_gr_dt",
                "transporter",
                "pack_type",
                "pack_mat",
                "mode_of_transport",
                "freight_term",
                "sec_freight_term",
                "delivery_id",
                "token_no",
                "di_date",
                "ship_city",
                "ship_taluka",
                "ship_district",
                "ship_state",
                "brand",
                "consignee",
                "customer_name",
                "cust_sub_category",
                "cust_category",
            )
        )

        merge_df = pd.merge(result, third_df, on="delivery_id", how="left")
        merge_df = merge_df.fillna("")

        merge_df["truck_reach_time"] = pd.to_datetime(
            merge_df["truck_reach_time"]
        ).dt.strftime("%d-%m-%Y %H:%M:%S")
        merge_df["truck_reach_time_confirm"] = pd.to_datetime(
            merge_df["truck_reach_time_confirm"]
        ).dt.strftime("%d-%m-%Y %H:%M:%S")
        merge_df["epod_completion_time"] = pd.to_datetime(
            merge_df["epod_completion_time"]
        ).dt.strftime("%d-%m-%Y %H:%M:%S")

        merge_df = merge_df.fillna("")
        merge_df = merge_df.to_dict(orient="records")

        return self.get_paginated_response(self.paginate_queryset(merge_df))


class TgtPlantDispatchDataDropdownAPIViewnew(GenericAPIView):
    queryset = TgtPlantDispatchData.objects.exclude(product="CLINKER")
    filter_backends = (DjangoFilterBackend,)
    # filterset_fields = ("di_so", "mode_of_transport")
    filterset_class = TgtPlantDispatchDataFilternew

    def get(self, request, *args, **kwargs):
        return Response(
            {
                "excise_invoice_no": self.filter_queryset(self.get_queryset())
                .values_list("excise_invoice_no", flat=True)
                .annotate(Count("excise_invoice_no"))
            }
        )


class SidingConstraintsCheckViewSet(GenericAPIView):
    queryset = SidingConstraints.objects.all()
    filter_backends = (DjangoFilterBackend,)

    def get(self, request, *args, **kwargs):
        rake_id = request.query_params.get("rake_id")
        if self.queryset.filter(rake_id=rake_id).exists():
            return Responses.success_response("data exist", data=True)
        else:
            return Responses.success_response("data not exist", data=False)


class HandlingMastersViewSet(ModelViewSet):
    queryset = HandlingMasters.objects.all()
    serializer_class = HandlingMastersSerializer
    filter_backends = (DjangoFilterBackend,)
    filterset_fields = ("rake_point", "rake_point_code", "depot_name", "packaging")
    # search_fields = ("code", "product")
    permission_classes = (IsHandlingAgent,)
    pagination_class = CustomPagination
    lookup_field = "id"

    def get_queryset(self):
        user = self.request.user
        user_roles = user.roles.values_list("role_name", flat=True).distinct()
        if UserRoleChoice.HA in user_roles:
            return (
                super()
                .get_queryset()
                .filter(depot_name__in=settings.USER_DEPOT_MAPPING.get(user.email))
            )
        return super().get_queryset()

    def get(self, request, *args, **kwargs):
        currentMonth = datetime.now().month
        currentyear = datetime.now().year

        queryset = self.filter_queryset(self.get_queryset())
        freight_master_queryset = DepoWiseFreightMaster.objects.all()

        if not queryset.exists():
            return self.get_paginated_response(self.paginate_queryset([]))

        handling_master_df = pd.DataFrame(
            queryset.values(
                "rake_point", "rake_point_code", "depot_name", "packaging"
            ).annotate(
                Avg("rake_handling"),
                Avg("cartage_cost"),
                Avg("road_handling_from_godown"),
            )
        )
        handling_master_df["depot_name_temp"] = handling_master_df["depot_name"].str[:3]
        if handling_master_df.empty:
            return self.get_paginated_response(self.paginate_queryset([]))
        # avg_freight_df = pd.DataFrame(
        #     DepoWiseFreightMaster.objects.values("pack_mat", "depo_code").annotate(
        #         Avg("average_freight")
        #     )
        # ).rename(columns={"depo_code": "depot_name", "pack_mat": "packaging"})

        avg_freight_df = pd.DataFrame(
            freight_master_queryset.filter(
                month_year__month=currentMonth, month_year__year=currentyear
            )
            .values("pack_mat", "depo_code")
            .annotate(Avg("average_freight"))
        ).rename(columns={"depo_code": "depot_name_temp", "pack_mat": "packaging"})

        if currentMonth == 1 and avg_freight_df.empty:
            avg_freight_df = pd.DataFrame(
                freight_master_queryset.filter(
                    month_year__month=12, month_year__year=currentyear - 1
                )
                .values("pack_mat", "depo_code")
                .annotate(Avg("average_freight"))
            ).rename(columns={"depo_code": "depot_name_temp", "pack_mat": "packaging"})

        if avg_freight_df.empty:
            avg_freight_df = pd.DataFrame(
                freight_master_queryset.filter(
                    month_year__month=currentMonth - 1, month_year__year=currentyear
                )
                .values("pack_mat", "depo_code")
                .annotate(Avg("average_freight"))
            ).rename(columns={"depo_code": "depot_name_temp", "pack_mat": "packaging"})
            if avg_freight_df.empty:
                avg_freight_df = pd.DataFrame(
                    columns=["packaging", "depot_name_temp", "average_freight__avg"]
                )

        merge_df = pd.merge(
            handling_master_df,
            avg_freight_df,
            on=["depot_name_temp", "packaging"],
            how="left",
        ).fillna("")
        merge_df.drop(["depot_name_temp"], axis=1, inplace=True)

        agent_name = pd.DataFrame(
            SidingWiseLiasioningAgent.objects.filter(
                siding_name__in=merge_df["rake_point"],
                siding_code__in=merge_df["rake_point_code"],
            ).values("liasioning_agent", "siding_name", "siding_code")
        ).rename(
            columns={"siding_name": "rake_point", "siding_code": "rake_point_code"}
        )
        if agent_name.empty:
            agent_name = pd.DataFrame(
                columns=["liasioning_agent", "rake_point", "rake_point_code"]
            )
        final_df = pd.merge(
            merge_df,
            agent_name,
            on=["rake_point", "rake_point_code"],
            how="left",
        ).fillna("")

        data = final_df.to_dict(orient="records")
        return self.get_paginated_response(self.paginate_queryset(data))


class GdVsWharfageModelRunView(GenericAPIView):
    """View class to create entries in LpModelDfFnl"""

    serializer_class = GdVsWharfageInputSerializer
    helper = GdVsWharfageHelper

    @transaction.atomic()
    def post(self, request, *args, **kwargs):
        """Insert data into GdVsWhargfage."""

        model_run_data = {
            "run_status": "Running",
            "rake_id": request.data.get("rake_id"),
            "rake_point": request.data.get("rake_point"),
            "rake_point_code": request.data.get("rake_point_code"),
            "next_rake_arrival_datetime": request.data.get(
                "next_rake_arrival_datetime"
            ),
            "wagon_quantity": request.data.get("wagon_quantity"),
            "run_date": date.today(),
        }
        model_run_serializer = GdVsWharfageInputSerializer(data=model_run_data)
        model_run_serializer.is_valid(raise_exception=True)
        model_run_instance = model_run_serializer.save()
        final_output, status = GdVsWharfageHelper.run_model(
            request.data.get("rake_id"),
            request.data.get("next_rake_arrival_datetime"),
            request.data.get("wagon_quantity"),
            request.data.get("rake_point_code"),
            request.data.get("serve_gd"),
            request.data.get("serve_crwc"),
            request.data.get("serve_wharfage"),
            request.data.get("operation_start"),
            request.data.get("operation_end"),
        )
        final_output.columns = final_output.columns.str.lower()
        final_output["crwc_freight_per_mt"] = final_output["crwc_freight_per_mt"].round(
            decimals=2
        )
        final_output["max_lifting_qty"] = final_output["max_lifting_qty"].round(
            decimals=2
        )
        final_output["gd_freight_per_mt"] = final_output["gd_freight_per_mt"].round(
            decimals=2
        )
        final_output["total_demurrage_freight_cost"] = final_output[
            "total_demurrage_freight_cost"
        ].round(decimals=2)
        final_output["total_wharfage_freight_cost"] = final_output[
            "total_wharfage_freight_cost"
        ].round(decimals=2)
        final_output["cost_demurrage"] = final_output["cost_demurrage"].round(
            decimals=2
        )
        final_output["cost_wharfage"] = final_output["cost_wharfage"].round(decimals=2)
        final_output["total_cost"] = final_output["total_cost"].round(decimals=2)
        final_output["total_crwc_freight_cost"] = final_output[
            "total_crwc_freight_cost"
        ].round(decimals=2)
        final_output["cost_crwc"] = final_output["cost_crwc"].round(decimals=2)
        shift_details_data = json.loads(final_output.to_json(orient="records"))

        gdvswharfage_details_serializer = GdVsWhargfageOutputSerializer(
            data=shift_details_data,
            context={
                "run_id": model_run_instance.run_id,
                "request_user": request.user.id,
            },
            many=True,
        )
        gdvswharfage_details_serializer.is_valid(raise_exception=True)
        gdvswharfage_details_serializer.save()
        model_run_instance.run_status = status
        model_run_instance.save()
        return Responses.success_response(
            "Successfully Running Data", data=model_run_instance.run_id
        )


class GdWharfageOutputViewSet(ModelViewSet):
    queryset = GdWharfageOutput.objects.all()

    def get(self, request, *args, **kwargs):
        run_id = request.query_params.get("run_id")
        queryset = self.queryset.filter(run_id=run_id)
        gd_wharfage_output_serializer = GdVsWharfageModelRunOutputByRunidSerializer(
            queryset, many=True, context={"run_id": run_id}
        )
        return Responses.success_response(
            "data fetched by id successfully", data=gd_wharfage_output_serializer.data
        )


class GetSlabsCode(GenericAPIView):
    def get(self, request):
        cnxn = connect_db()
        next_rake_arrival_datetime = request.query_params.get("next_arrival_date")
        rake_id = request.query_params.get("rake_id")
        rake_point = request.query_params.get("rake_point")
        rake_details = pd.read_sql(
            f"""SELECT "RAKE_ID",
            "RAKE_POINT",
            "RAKE_POINT_CODE",
            "DEPOT_CODE",
            "QTY",
            "BRAND",
            "GRADE",
            "PACKAGING",
            "CUST_CATEGORY",
            "PLACEMENT_TIME",
            "RELEASE_TIME",
            "DEMURRAGE_START",
            "WHARFAGE_START"
            FROM
            (select "RLD_ID",
            "PLACEMENT_DATE" as "PLACEMENT_TIME",
            "ACTUAL_TIME_FOR_RAKE_RELEASE" as "RELEASE_TIME",
            "FREE_TIME_FOR_RAKE_RELEASE" as "DEMURRAGE_START",
            "FREE_TIME_FOR_MATERIAL_CLEARANCE" as "WHARFAGE_START"
            from target."TGT_RAKE_UNLOADING_DETAILS") A,
            (select rld."RLD_ID",
            rld."RAKE_ID",
            rld."RAKE_POINT",
            rld."RAKE_POINT_CODE",
            SUBSTR(rld."SHIP_TO_DEPOT",1,3) AS "DEPOT_CODE",
            rld."QTY_DISPATCH_FRM_PLANT" AS "QTY",
            rld."ORG_ID" AS "BRAND",
            rld."INVENTORY_ITEM_ID",
            case
            WHEN rld."PACKING_TYPE" = 'PAPER' then 'PAPER'
                        else 'HDPE'
                    end as "PACKAGING",
            rld."SEGMENT" AS "CUST_CATEGORY",a."MEANING" as "GRADE" from target."TGT_RAKE_LOADING_DETAILS" rld,
            (select "MEANING","LOOKUP_CODE" from etl_zone."T_OEBS_FND_LOOKUP_VALUES" oflv
            where oflv."LOOKUP_TYPE" = 'SCL_CEMENT_ITEMS'
            and oflv."ZD_EDITION_NAME" = 'SET2') a
            where rld."INVENTORY_ITEM_ID"::bigint = a."LOOKUP_CODE"::bigint)B

            WHERE A."RLD_ID" = B."RLD_ID" AND  "RAKE_ID"='{rake_id}' and "RAKE_POINT"='{rake_point}'
            """,
            cnxn,
        )
        slabs = pd.DataFrame(columns=["RAKE_ID", "START_TIME", "END_TIME"])
        slabs.loc[len(slabs)] = rake_details.iloc[0][
            ["RAKE_ID", "PLACEMENT_TIME", "DEMURRAGE_START"]
        ].tolist()
        slabs.loc[len(slabs)] = rake_details.iloc[0][
            ["RAKE_ID", "DEMURRAGE_START", "WHARFAGE_START"]
        ].tolist()
        if slabs.iloc[-1]["END_TIME"].hour != 0:
            slabs.loc[len(slabs)] = slabs.iloc[-1][["RAKE_ID", "END_TIME"]].tolist() + [
                (slabs.iloc[-1]["END_TIME"] + td(days=1)).replace(
                    hour=0, minute=0, second=0
                )
            ]
        for date_ in pd.date_range(
            start=slabs.iloc[-1]["END_TIME"] + td(days=1),
            end=next_rake_arrival_datetime,
            freq="d",
        ):
            slabs.loc[len(slabs)] = slabs.iloc[-1][["RAKE_ID", "END_TIME"]].tolist() + [
                date_
            ]
        slabs["size"] = 1
        bgpt = rake_details.groupby(
            ["BRAND", "GRADE", "PACKAGING", "CUST_CATEGORY"], as_index=False
        ).size()
        bgpt["size"] = 1
        slabs = slabs.merge(bgpt, on=["size"])
        slabs.drop(columns=["size"], inplace=True)
        slabs["SLAB_ID"] = slabs.index + 1
        slabs = slabs[["START_TIME", "END_TIME"]].drop_duplicates()
        slabs["Time_Slabs"] = (
            slabs["START_TIME"].astype(str) + "  " + slabs["END_TIME"].astype(str)
        )
        slabs = slabs.drop(
            [
                # "BRAND",
                # "GRADE",
                # "PACKAGING",
                # "CUST_CATEGORY",
                # "SLAB_ID",
                # "RAKE_ID",
                "START_TIME",
                "END_TIME",
            ],
            axis=1,
        )
        return Responses.success_response("getting slabs successfully", data=slabs)


class FreightChangeFormatDownloadView(ModelViewSet):
    def get(self, request, *args, **kwargs):
        with open(
            f"analytical_data/DS_model/freight_change_initiation.xlsx",
            "rb",
        ) as excel:
            workbook = excel.read()

        content_type = (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response = HttpResponse(workbook, content_type=content_type)
        response[
            "Content-Disposition"
        ] = f"attachment; filename=freight_change_initiation.xlsx"

        return response
